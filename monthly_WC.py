'''
This script generates a report to track the performance of each supplier,
and their daily inventory, payable, COGS and inbounds based on an excel template
'''

import os
import argparse
import subprocess
import sys
import datetime as dt
import pandas as pd
import dateutil.relativedelta
import openpyxl
from openpyxl.utils import get_column_letter
from pyspark.sql import SparkSession

reload(sys)
sys.setdefaultencoding('utf8')

SPARK = SparkSession \
    .builder \
    .enableHiveSupport() \
    .getOrCreate()
SPARK.sql('use shopee')


def get_sop_eop(current_month):
    '''
    This function returns the SOP (Start of Period) and EOP (End of Period) for the
    previous month which is denoted as the first day of the last month (SOP) and the
    last day of the last month (EOP) respectively.
    :return SOP, EOP: start of period, end of period
    '''
    two_months_back = current_month - dateutil.relativedelta.relativedelta(months=2)
    m6_sop = two_months_back.replace(day=1)
    return m6_sop

def get_days_in_months(today_date):
    '''
    Returns the number of days in each month assuming there are only three months
    The last number will always be 30 if today is not the end of the month
    :return: days_in_month (a list of number of days in each month)
    '''
    start_date = get_sop_eop(today_date)
    days_in_months = [0,
                      pd.Period(start_date.strftime("%Y-%m-%d")).days_in_month,
                      pd.Period((start_date + pd.offsets.MonthBegin(1)).
                                strftime("%Y-%m-%d")).days_in_month,
                      30]
    return days_in_months


def append_to_excel(write_dict, supplier_info_cols, today_date, days_in_months, wb_obj):
    '''
    This function paste the values in specified tables into the template,
    and automatically computes the averages and sums of one month
    :param write_dict: a dictionary of pivot tables
    :param supplier_info_cols: the columns of supplier category and names
    :param start_date: start of period
    :param days_in_months: number of days in each month
    :param wb_obj: excel template to be written
    :return:
    '''
    start_row = 3
    start_date = get_sop_eop(today_date)
    for sheet_name, pivot_df in write_dict.items():
        work_sheet = wb_obj[sheet_name]
        work_sheet['F2'] = start_date
        for df_index, row in pivot_df.iterrows():
            row_index = df_index + start_row
            for df_col_name, excel_letter in supplier_info_cols.items():
                cell_index = excel_letter + str(row_index)
                work_sheet[cell_index] = row[df_col_name]
            for column in range(6, pivot_df.shape[1] + 4):
                day_index = column - 4
                column_letter = get_column_letter(column)
                cell_index = column_letter + str(row_index)
                work_sheet[cell_index] = row[day_index]
            for column in range(3, 6):
                column_letter = get_column_letter(column)
                month_index = column - 2
                if month_index < len(days_in_months) - 1:
                    start_col = 2 + sum(days_in_months[:month_index])
                    end_col = 2 + sum(days_in_months[:month_index + 1])
                    cell_index = column_letter + str(row_index)
                    if sheet_name in ['Daily Accounts Payable', 'Daily Inventory Value']:
                        work_sheet[cell_index] = row[start_col:end_col].mean()
                    else:
                        work_sheet[cell_index] = row[start_col:end_col].sum()
                elif month_index == len(days_in_months) - 1:
                    end_col = 2 + sum(days_in_months[:month_index + 1])
                    start_col = end_col - 30
                    cell_index = column_letter + str(row_index)
                    if sheet_name in ['Daily Accounts Payable', 'Daily Inventory Value']:
                        work_sheet[cell_index] = row[start_col:end_col].mean()
                    else:
                        work_sheet[cell_index] = row[start_col:end_col].sum()
                else:
                    work_sheet[cell_index] = float('nan')
            for column in range(pivot_df.shape[1] + 4, work_sheet.max_column):
                column_letter = get_column_letter(column)
                cell_index = column_letter + str(row_index)
                work_sheet[cell_index] = None


def get_main_df(country, today_date):
    '''
    Get the data for the last three months for specific country
    '''
    m6_sop = get_sop_eop(today_date).strftime("%Y-%m-%d")
    today_date = today_date.strftime("%Y-%m-%d")
    query = '''
    SELECT category_cluster, supplier_name, sku_id, sourcing_status, 
            cogs_usd, stock_on_hand, inbound_value_usd, acct_payables_usd, 
            inventory_value_usd, payment_terms, brand, grass_date 
    from shopee_bi_sbs_mart
    where
        grass_date >= date('{start_date}')
    and
        grass_date <= date('{end_date}')
    AND
        purchase_type = 'Outright'
    and 
        grass_region = '{cntry}'
    '''.format(start_date=m6_sop, end_date=today_date, cntry=country)
    print(query)
    return SPARK.sql(query)


def write_to_csv(query_df, report_directory_name, output_file_name):
    '''
    This function writes the spark dataframe to a csv file
    :param query_df: spark dataframe
    :type query_df: spark dataframe
    :param output_file_name: Name of Output File
    :type output_file_name: str
    :param report_directory_name: Name of Report Directory
    :type report_directory_name: str
    :return:
    '''
    print("Output File Name: ", output_file_name)
    req_headers = query_df.columns
    output_filename = '/user/yutong.zou/{}/tmp.csv'.format(report_directory_name)
    local_filename = output_file_name
    query_df.write.format('csv').mode('overwrite'). \
        options(header='false', escape='"', encoding="UTF-8").save(output_filename)
    subprocess.call('/usr/share/hadoop/bin/hadoop fs -getmerge %s %s' % (
        output_filename, local_filename + "_tmp"), shell=True)
    subprocess.call('/usr/share/hadoop/bin/hadoop fs -rmr %s' % output_filename, shell=True)
    db_input_file = open(local_filename + "_tmp", 'rb')
    print(db_input_file)
    db_output_file = open(local_filename, 'wb')
    db_output_file.write(','.join(req_headers) + '\n')
    for line in db_input_file:
        db_output_file.write(line)
    db_output_file.close()
    os.remove(local_filename + "_tmp")


def main(today_date, countries, data_queried):
    for country in countries:
        print('starting country', country)
        country_folder_path = './' + country + '/'

        if not os.path.exists(country_folder_path):
            os.makedirs(country_folder_path)

        input_file_path = "{}input_files".format(country_folder_path)

        if not os.path.exists(input_file_path):
            os.makedirs(input_file_path)

        if not data_queried:
            df = get_main_df(country, today_date)
            write_to_csv(df, 'main_df', input_file_path + '/main_df.csv')
            df = df.toPandas()
        else:
            df = pd.read_csv(input_file_path + '/main_df.csv', encoding='utf-8-sig', index_col=False)

        df['grass_date'] = pd.to_datetime(df['grass_date'])
        df_info = df[df['grass_date'] == today_date]

        df_inv_count = df_info.groupby('supplier_name')['stock_on_hand'].sum().reset_index()
        df_inv_count.columns = ['supplier_name', 'inv_count']

        df_inv_sorting = df_info.groupby(['category_cluster', 'supplier_name'])[
            'inventory_value_usd'].sum().reset_index()
        df_inv_sorting0 = df_inv_sorting.sort_values(['supplier_name', 'inventory_value_usd'], ascending=False)
        df_inv_sorting1 = df_inv_sorting0.groupby('supplier_name')['supplier_name', 'category_cluster'].head(1)
        df_inv_sorting1 = df_inv_sorting1[['category_cluster', 'supplier_name']]
        df_inv_count = df_inv_sorting1.merge(df_inv_count, on=['supplier_name'], how='left')
        df_inv_sum = df_info.groupby('supplier_name')['inventory_value_usd',].sum().reset_index()

        df_sku_count = df_info.groupby('supplier_name')['sku_id'].count().reset_index()
        df_sku_count.columns = ['supplier_name', 'no_skus_WH']
        df_info0 = df_sku_count.merge(df_inv_count, on=['supplier_name'], how='left')

        df_payment = df_info.groupby('supplier_name')['supplier_name', 'payment_terms'].head(1).reset_index(drop=True)
        df_last_month = df[(df['grass_date'] >= today_date - dt.timedelta(days=30)) & (df['grass_date'] <= today_date)]
        df_last_month.drop_duplicates(inplace=True)
        brands_df = df_last_month[['supplier_name', 'brand']]
        brands_df = brands_df[brands_df.brand != 'nan']
        number_of_top_brands = 3
        supplier_names = set(brands_df['supplier_name'].to_list())
        brand_count_df = brands_df.groupby(['supplier_name', 'brand']).size()
        # get nlargest per subgroup
        brand_count_top_3_df = brand_count_df.groupby(['supplier_name'], group_keys=False).apply(
            lambda subgroup: subgroup.nlargest(3))
        # drop count column and reset_index
        brand_count_top_3_df = brand_count_top_3_df.reset_index().drop(columns=0)
        # final brands dataframe in desired output
        top_brands_cols = ['supplier_name', 'brand_1', 'brand_2', 'brand_3']
        top_brands_df = pd.DataFrame(columns=top_brands_cols)
        for supplier in supplier_names:
            slice_df = brand_count_top_3_df[(brand_count_top_3_df.supplier_name == supplier)].reset_index()
            three_brands = dict()
            for i in range(number_of_top_brands):
                key = 'brand_{}'.format(i + 1)
                try:
                    brand = slice_df.iloc[i].brand
                    three_brands[key] = brand
                except IndexError:
                    three_brands[key] = 'n.a.'
            row_list = [supplier, three_brands['brand_1'], three_brands['brand_2'],
                        three_brands['brand_3']]
            row_to_append = pd.DataFrame([row_list], columns=top_brands_cols)
            top_brands_df = top_brands_df.append(row_to_append)

        top_brands_df = top_brands_df.reset_index(drop=True)
        top_brands_df.fillna('n.a.', inplace=True)
        df_info1 = df_info0.merge(df_inv_sum, on=['supplier_name'], how='left')
        df_info2 = df_info1.merge(top_brands_df, on=['supplier_name'], how='left')
        df_info3 = df_info2.merge(df_payment, on=['supplier_name'], how='left')
        df_total_sum = df.groupby(['supplier_name', 'grass_date']).sum().reset_index()
        df_total_sum['grass_date'] = df_total_sum['grass_date'].dt.strftime('%Y-%m-%d')
        df_cogs = pd.pivot_table(df_total_sum, values='cogs_usd', index='supplier_name',
                                 columns='grass_date').reset_index()
        df_cogs = df_inv_sorting1.merge(df_cogs, on=['supplier_name'], how='right')

        df_payables = pd.pivot_table(df_total_sum, values='acct_payables_usd', index='supplier_name',
                                     columns='grass_date').reset_index()
        df_payables = df_inv_sorting1.merge(df_payables, on=['supplier_name'], how='right')

        df_inv_value = pd.pivot_table(df_total_sum, values='inventory_value_usd', index='supplier_name',
                                      columns='grass_date').reset_index()
        df_inv_value = df_inv_sorting1.merge(df_inv_value, on=['supplier_name'], how='right')

        df_inbound = pd.pivot_table(df_total_sum, values='inbound_value_usd', index='supplier_name',
                                    columns='grass_date').reset_index()
        df_inbound = df_inv_sorting1.merge(df_inbound, on=['supplier_name'], how='right')

        df_info3 = df_info3[['category_cluster', 'supplier_name', 'no_skus_WH', 'inv_count',
                            'inventory_value_usd', 'brand_1', 'brand_2', 'brand_3',
                            'payment_terms']]

        path = country_folder_path + 'Weekly_wc_template.xlsx'
        wb_obj = openpyxl.load_workbook(path)
        work_sheet = wb_obj[country]
        m_row = work_sheet.max_row
        supplier_highlight = []
        for i in range(4, m_row + 1):
            cell_obj = work_sheet.cell(row=i, column=1)
            if cell_obj.value in (None, 'TOTAL'):
                continue
            supplier_highlight.append(cell_obj.value)

        df_info4 = df_info3[df_info3['supplier_name'].isin(supplier_highlight)].reset_index(drop=True)
        main_ws = wb_obj['Tracking']
        main_ws['B1'] = dt.date.today()
        for df_index, row in df_info4.iterrows():
            for col_index in range(df_info4.shape[1]):
                cell_index = get_column_letter(col_index + 1) + str(df_index + 4)
                main_ws[cell_index] = row[col_index]

        write_dict = {
            'Daily Accounts Payable': df_payables,
            'Daily Inventory Value': df_inv_value,
            'Daily COGS': df_cogs,
            'Daily Inbounds': df_inbound
        }
        supplier_info_cols = {
            "category_cluster": "A",
            "supplier_name": "B",
        }
        days_in_months = get_days_in_months(today_date)
        append_to_excel(write_dict, supplier_info_cols, today_date, days_in_months, wb_obj)
        wb_obj.save(country_folder_path + '/month_{}_sample.xlsx'.format(country))

        df_info3.to_csv(country_folder_path + '{}_tracking_tab_v2.csv'.format(country), encoding="utf-8-sig")
        df_payables.to_csv(country_folder_path + '{}_payables_v2.csv'.format(country), encoding="utf-8-sig")
        df_cogs.to_csv(country_folder_path + '{}_cogs_v2.csv'.format(country), encoding="utf-8-sig")
        df_inv_value.to_csv(country_folder_path + '{}_inventory_value_v2.csv'.format(country), encoding="utf-8-sig")
        df_inbound.to_csv(country_folder_path + '{}_inbound_v2.csv'.format(country), encoding="utf-8-sig")


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('-c', '--countries', nargs='+', default=['ID', 'MY', 'TH', 'VN', 'TW', 'PH'])
    parser.add_argument("-d", "--date",
                        default=dt.date.today(),
                        type=lambda d: dt.datetime.strptime(d, '%Y%m%d').date(),
                        help="Date in the format yyyymmdd")
    parser.add_argument('-q', '--queried', help='has data been queried?', default=False)
    args = parser.parse_args()
    main(args.date, args.countries, args.queried)
