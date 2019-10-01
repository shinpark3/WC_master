'''
This script generates a report to track the performance of each supplier,
and their daily inventory, payable, COGS and inbounds based on an excel template
'''

import os
import argparse
import calendar
import subprocess
import string
import sys
from collections import Counter
import datetime as dt
import pandas as pd
import numpy as np
import dateutil.relativedelta
import openpyxl
from openpyxl.utils import get_column_letter
from pyspark.sql.functions import col
from pyspark.sql import SparkSession

reload(sys)
sys.setdefaultencoding('utf8')

SPARK = SparkSession \
    .builder \
    .enableHiveSupport() \
    .getOrCreate()
SPARK.sql('use shopee')


def get_usd_exchange_rate(country, eop):
    '''
    This function returns the exchange rate on specific date for the specified country
    :return: rate: exchange rate
    '''
    eop = eop.strftime("%Y-%m-%d")
    query = '''
    SELECT 
        exchange_rate 
    from 
        dim_exrate 
    where 
        grass_date = date('{end_period}') 
    and 
        country = '{cntry}'
    '''.format(end_period=eop, cntry=country)
    rate = SPARK.sql(query).head()[0]
    return float(rate)


def get_sop_eop(current_month):
    '''
    This function returns the SOP (Start of Period) and EOP (End of Period) for the
    previous month which is denoted as the first day of the last month (SOP) and the
    last day of the last month (EOP) respectively.
    :return SOP, EOP: start of period, end of period
    '''
    two_months_back = current_month - dateutil.relativedelta.relativedelta(months=2)
    # three_months_back = current_month - dateutil.relativedelta.relativedelta(months=3)
    # six_months_back = current_month - dateutil.relativedelta.relativedelta(months=6)
    m1_end_of_month_day = calendar.monthrange(current_month.year, current_month.month)[1]
    # m2_end_of_month_day = calendar.monthrange(two_months_back.year, two_months_back.month)[1]
    m1_sop = current_month.replace(day=1)
    m1_eop = current_month.replace(day=m1_end_of_month_day)
    m6_sop = two_months_back.replace(day=1)
    # m6_eop = two_months_back.replace(day=m2_end_of_month_day)
    return m1_sop, m1_eop, m6_sop


def get_latest_cogs(cogs_df):
    '''
    Returns the latest cost of goods sold (COGS) for each sku in cogs_df
    :return: cogs3: COGS for each sku
    '''
    cogs0 = cogs_df.sort_values('grass_date', ascending=False)
    cogs1 = cogs0[['sku_id', 'stored_value_fifo_per_qty', 'grass_date']]
    cogs2 = cogs1[(cogs1['stored_value_fifo_per_qty'] != 0) &
                  (~cogs1['stored_value_fifo_per_qty'].isnull())]
    cogs3 = cogs2.groupby('sku_id').head(1)
    return cogs3


def get_sop(current_month):
    '''
    Returns the first day of the month
    :return: first day of current month
    '''
    return current_month.replace(day=1)


def get_wms_info(eop, country):
    '''
    Queries the database for wms_info within 150 days before eop
    :return: wms_info: outright info
    '''
    print('wms_info at latest EOP date')
    path = '/user/shopeebi/business_analytics/sbs/mart/sku_details'
    print("Outright info ==> ", eop.strftime("%Y-%m-%d"))
    wms_info = SPARK.read.parquet(path).filter(
        col('purchase_type') == 'Outright').distinct()  # REMOVE distinct next time
    wms_info = wms_info.filter((col('grass_date')
                                >= '%s' % (eop - dt.timedelta(days=150)).strftime("%Y-%m-%d"))
                               & (col('grass_region') == country))
    return wms_info


def get_pr_po(country):
    '''
    Queries database for pr_po records in one country
    :return: pr_po table
    '''
    query = '''
    select 
        pr_sku.sku_id, 
        pr_sku.reason as pr_reason, 
        pr_sku.quantity as pr_quantity, 
        date(from_unixtime(pr_tab.ctime)) as pr_date, 
        pr_tab.country, 
        pr_tab.request_id as pr_id, 
        pr_sku.sourcing_status, 
        pr_tab.status as pr_status, 
        pr_tab.status_text, 
        po_hourly.order_id,
        po_hourly.status as po_status,
        date(from_unixtime(po_hourly.ctime)) as po_date,
        po_sku.confirmed_quantity as po_quantity,
        date(from_unixtime(inbound_tab.inbound_time)) as inbound_time, 
        inbound_tab.quantity as inbound_quantity,
        inbound_tab.inbound_id as inbound_id
    from shopee_pms_db__purchase_request_tab as pr_tab
    left join shopee_pms_db__purchase_request_sku_tab as pr_sku
        on pr_tab.id = pr_sku.pr_id
    left join shopee_pms_db__purchase_order_tab_hourly as po_hourly
        on pr_tab.request_id = po_hourly.request_id
    left join shopee_pms_db__purchase_order_sku_tab as po_sku
        on po_hourly.order_id = po_sku.order_id and pr_sku.sku_id = po_sku.sku_id
    left join shopee_pms_db__purchase_order_inbound_sku_tab as inbound_tab
        on inbound_tab.order_id = po_hourly.order_id and pr_sku.sku_id = inbound_tab.sku_id
    where pr_sku.reason != 'Re-Inbound'
    and pr_tab.country = '{}'
    '''.format(country)
    print(query)
    return SPARK.sql(query)


def get_color(sop, country):
    '''
    Queries the inventory database after sop (starting date) of one country
    :return: inventory table
    '''
    sop = sop.strftime("%Y-%m-%d")
    query = '''
    select sku_id, inventory_value_usd, grass_date, color, stock_on_hand 
    from shopee_bi_wms_color 
    where grass_region = '{cntry}' 
    and grass_date >= date('{starting_date}') 
    and purchase_type = 'Outright'
    '''.format(cntry=country, starting_date=sop)
    print(query)
    return SPARK.sql(query)


def get_payables(ending_payable_date):
    '''
    Queries database for accounts payable information on specific date
    :return: accounts payable for each sku on specified date
    '''
    all_LP = SPARK.read.parquet('/projects/regbida_ba/sbs/mart/6a_payables') \
        .filter(col('grass_date') == '%s' % (ending_payable_date)) \
        .select('sku_id', 'end_payables').distinct()
    return all_LP


def get_cogs(country):
    '''
    Queries database for COGS (cost of goods sold) information of the country
    :return: COGS of each sku in the country
    '''
    query = '''
    select sku_id, stored_value_fifo_per_qty, grass_date 
    from shopee_bi_sku_cogs 
    where grass_region = '%c'
    '''
    query = query.replace('%c', str(country))
    print(query)
    return SPARK.sql(query)


def get_replenishment(country):
    '''
    Queries database for replenishment data of the country within 150 days
    :return: replenishment data
    '''
    query = '''
    SELECT sku_id, sales, flash_sales, (sales + flash_sales) as total_sales, grass_date 
    from shopee_bi_replenishment_data_pool_core 
    where contract_type in ('B2C') 
    and country = '{cntry}' 
    and grass_date >= current_date - interval '150' day'''.format(cntry=country)
    print(query)
    return SPARK.sql(query)


def get_purchases(sop, eop, country):
    '''
    Queries the database of the country between the specified start and end date
    :return: purchase info
    '''
    query = '''
    with pr as (
            select 
                sku_id,
                date(from_unixtime(ctime)) as pr_created_date,
                request_id
            from shopee_pms_db__purchase_request_tab a
            join shopee_pms_db__purchase_request_sku_tab c on
                a.id = c.pr_id
            where 
                date(from_unixtime(ctime)) <= date('{end_date}')
            and 
                date(from_unixtime(ctime)) >= date('{start_date}')
            AND 
                status in (4)
            AND 
                reason != 'Re-Inbound'
            AND 
                country = '{ctry}'
        ), po as (
            select
                po.request_id,
                sku_id,
                value  as po_value
            from shopee_pms_db__purchase_order_tab_hourly po
            join shopee_pms_db__purchase_order_sku_tab pos ON
                po.order_id = pos.order_id
            where po.status in (1, 2, 3, 5)
        )
        select 
            pr.sku_id,
            po.po_value as po_value,
            pr.pr_created_date
        from pr 
        join po ON
            pr.request_id = po.request_id 
            and
            pr.sku_id = po.sku_id'''.format(ctry=country, start_date=sop, end_date=eop)
    print(query)
    return SPARK.sql(query)


def get_soh(eop):
    '''
    Return black inventory
    :param eop: End Period
    :return black_df: dataframe for black inventory
    '''
    eop = eop.strftime("%Y-%m-%d")
    query = '''
    SELECT DISTINCT
       sku_id,
       stock_on_hand as end_qty
    FROM
       shopee_bi_wms_color 
    WHERE
       purchase_type in ('Outright', 'Consignment', 'Fulfilment')
    AND 
        date(grass_date) = date('{end_period}')
    '''.format(end_period=eop)
    print(query)
    soh_df = SPARK.sql(query)
    return soh_df


def get_days_in_months(pivot_df):
    '''
    Returns the number of days in each month of the pivot tables with dates as column names
    Assumed there are only three months.
    The last number will always be 30 if today is not the end of the month
    :param pivot_df: cogs_pivot, inventory_pivot, payables_pivot or purchase_pivot
    :return: start_date (start date in the data)
             days_in_month (a list of number of days in each month)
    '''
    start_date = dt.datetime.strptime(pivot_df.columns[2], "%Y-%m-%d")
    end_date = dt.datetime.strptime(pivot_df.columns[-1], "%Y-%m-%d")
    # last_day_of_last_month = calendar.monthrange(end_date.year, end_date.month)[1]
    # if end_date.day == last_day_of_last_month:
    #     days_in_last_month = end_date.day
    # else:
    #     days_in_last_month = 30
    days_in_months = [0, pd.Period(start_date.strftime("%Y-%m-%d")).days_in_month,
                      pd.Period((start_date + pd.offsets.MonthBegin(1)).
                                strftime("%Y-%m-%d")).days_in_month,
                      end_date.day]
    return start_date, days_in_months


def write_to_csv(query_df, report_directory_name, output_file_name):
    '''
    This function writes the spark dataframe to a csv file
    :param query_df: spark dataframe
    :type query_df: spark dataframe
    :param output_file_name: Name of Output File
    :type output_file_name: str
    :param report_directory_name: Name of Report Directory
    :type report_directory_name: str
    :return:
    '''
    print("Output File Name: ", output_file_name)
    req_headers = query_df.columns
    output_filename = '/user/yutong.zou/{}/tmp.csv'.format(report_directory_name)
    local_filename = output_file_name
    query_df.write.format('csv').mode('overwrite'). \
        options(header='false', escape='"', encoding="UTF-8").save(output_filename)
    subprocess.call('/usr/share/hadoop/bin/hadoop fs -getmerge %s %s' % (
        output_filename, local_filename + "_tmp"), shell=True)
    subprocess.call('/usr/share/hadoop/bin/hadoop fs -rmr %s' % output_filename, shell=True)
    db_input_file = open(local_filename + "_tmp", 'rb')
    print(db_input_file)
    db_output_file = open(local_filename, 'wb')
    db_output_file.write(','.join(req_headers) + '\n')
    for line in db_input_file:
        db_output_file.write(line)
    db_output_file.close()
    os.remove(local_filename + "_tmp")


def append_to_excel(write_dict, supplier_info_cols, start_date, days_in_months, wb_obj):
    '''
    This function paste the values in specified tables into the template,
    and automatically computes the averages and sums of one month
    :param write_dict: a dictionary of pivot tables
    :param supplier_info_cols: the columns of supplier category and names
    :param start_date: start of period
    :param days_in_months: number of days in each month
    :param wb_obj: excel template to be written
    :return:
    '''
    start_row = 3
    for sheet_name, pivot_df in write_dict.items():
        work_sheet = wb_obj[sheet_name]
        work_sheet['F2'] = start_date
        for df_index, row in pivot_df.iterrows():
            row_index = df_index + start_row
            for df_col_name, excel_letter in supplier_info_cols.items():
                cell_index = excel_letter + str(row_index)
                work_sheet[cell_index] = row[df_col_name]
            for column in range(6, pivot_df.shape[1] + 4):
                day_index = column - 4
                column_letter = get_column_letter(column)
                cell_index = column_letter + str(row_index)
                work_sheet[cell_index] = row[day_index]
            for column in range(3, 6):
                column_letter = get_column_letter(column)
                month_index = column - 2
                if month_index < len(days_in_months) - 1:
                    start_col = 2 + sum(days_in_months[:month_index])
                    end_col = 2 + sum(days_in_months[:month_index + 1])
                    cell_index = column_letter + str(row_index)
                    if sheet_name in ['Daily Accounts Payable', 'Daily Inventory Value']:
                        work_sheet[cell_index] = row[start_col:end_col].mean()
                    else:
                        work_sheet[cell_index] = row[start_col:end_col].sum()
                elif month_index == len(days_in_months) - 1:
                    end_col = 2 + sum(days_in_months[:month_index + 1])
                    start_col = end_col - 30
                    cell_index = column_letter + str(row_index)
                    if sheet_name in ['Daily Accounts Payable', 'Daily Inventory Value']:
                        work_sheet[cell_index] = row[start_col:end_col].mean()
                    else:
                        work_sheet[cell_index] = row[start_col:end_col].sum()
                else:
                    work_sheet[cell_index] = float('nan')
            for column in range(pivot_df.shape[1] + 4, work_sheet.max_column):
                column_letter = get_column_letter(column)
                cell_index = column_letter + str(row_index)
                work_sheet[cell_index] = None


def get_clean_brands(brands_df):
    '''
    This function takes in a dataframe with brands and remaps all brands to clean brands.
    For example, if the data contains brands of "xiao mi" and "xiaomi",
    all brands will be renamed to "xiao mi"
    :param brands_df:
    :return brands_df: dataframe of clean brands
    '''
    before_brands = brands_df['brand'].unique()
    # if number of brands == 0, just return brands_df
    if before_brands.shape[0] == 0:
        return brands_df
    # vectorized ambda function to replace whitespace inside nparray
    replace = lambda x: string.replace(x, " ", "")
    vfunc = np.vectorize(replace)
    brands_removed_whitespace = vfunc(before_brands)
    count_of_brands = Counter(brands_removed_whitespace)
    # get brands who have count > 1 (number of version for brands more than 1)
    brands_more_than_1 = []
    for brand, counter in count_of_brands.items():
        if counter > 1:
            brands_more_than_1.append(brand)
    # make dictionary to remap brand names
    # for e.g "xiaomi" and "xiao mi" will now all be renamed to "xiao mi"
    before_brands = list(before_brands)
    brands_remap = {}
    for brand in before_brands:
        strip_space_brand = string.replace(brand, " ", "")
        if brand != strip_space_brand and strip_space_brand in brands_removed_whitespace \
                and strip_space_brand in brands_more_than_1:
            brands_remap[strip_space_brand] = brand
    brands_df['brand'] = brands_df['brand'].apply(
        lambda brand: brands_remap[brand] if brand in brands_remap.keys() else brand)
    return brands_df


def main(today_date, countries, data_queried):
    m1_sop, m1_eop, m6_sop = get_sop_eop(today_date)
    beg_payable = get_payables(m6_sop - dt.timedelta(days=1)).toPandas()
    beg_payable.to_csv('beg_payable.csv')
    for country in countries:
        print('starting country', country)
        country_folder_path = './' + country + '/'

        if not os.path.exists(country_folder_path):
            os.makedirs(country_folder_path)

        input_file_path = "{}input_files".format(country_folder_path)

        if not os.path.exists(input_file_path):
            os.makedirs(input_file_path)

        if not data_queried:
            print('1. Getting wms info')
            df = get_wms_info(today_date, country)
            print('2. Getting color')
            color = get_color(m6_sop, country)
            print('3. Getting pr po info')
            pr_po = get_pr_po(country)
            print('4. Getting cogs')
            cogs = get_cogs(country)
            print('5. Getting replenishment')
            replenishment = get_replenishment(country)
            print('6. Getting purchases')
            purchases = get_purchases(m6_sop, today_date, country)
            print('7. Getting ending stock on hand')
            soh_end = get_soh(m1_eop)
            write_to_csv(df, 'wms_info', input_file_path + '/wms_info.csv')
            write_to_csv(color, 'color', input_file_path + '/color.csv')
            write_to_csv(pr_po, 'pr_po', input_file_path + '/pr_po.csv')
            write_to_csv(cogs, 'cogs', input_file_path + '/cogs.csv')
            write_to_csv(replenishment, 'replenishment', input_file_path + '/replenishment.csv')
            write_to_csv(purchases, 'purchases', input_file_path + '/purchases.csv')
            write_to_csv(soh_end, 'soh_end', input_file_path + '/soh_end.csv')
            print('8. Finished writing all queries from database to CSV')
            df = df.toPandas()
            color = color.toPandas()
            pr_po = pr_po.toPandas()
            cogs = cogs.toPandas()
            replenishment = replenishment.toPandas()
            purchases = purchases.toPandas()
            print('9. Finished converting all dataframes to Pandas Dataframe')
        else:
            print('1. Starting to read all data')
            df = pd.read_csv(input_file_path + '/wms_info.csv', encoding='utf-8-sig', index_col=False)
            color = pd.read_csv(input_file_path + '/color.csv', encoding='utf-8-sig', index_col=False)
            pr_po = pd.read_csv(input_file_path + '/pr_po.csv', encoding='utf-8-sig', index_col=False)
            cogs = pd.read_csv(input_file_path + '/cogs.csv', encoding='utf-8-sig', index_col=False)
            purchases = pd.read_csv(input_file_path + '/purchases.csv',
                                    encoding='utf-8-sig', index_col=False)
            replenishment = pd.read_csv(input_file_path + '/replenishment.csv',
                                        encoding='utf-8-sig', index_col=False)
            print('2. Finished reading all data')

        replenishment['grass_date'] = pd.to_datetime(replenishment['grass_date'])
        color['grass_date'] = pd.to_datetime(color['grass_date'])
        country_x_rate = get_usd_exchange_rate(country, today_date)
        df0 = df[['sku_id', 'category_cluster', 'purchase_type', 'brand',
                  'sourcing_status', 'supplier_id', 'supplier_name',
                  'payment_terms', 'payment_days', 'is_payment_afterme',
                  'grass_region', 'grass_date']]
        df0.dropna(inplace=True)
        df0['grass_date'] = pd.to_datetime(df0['grass_date'])
        df1 = df0[['sku_id', 'category_cluster', 'payment_days',
                   'grass_date', 'brand', 'supplier_name']]
        m6_df = df1[(df1['grass_date'] >= m6_sop) & (df1['grass_date'] <= today_date)]
        m6_df1 = df1[df1['grass_date'] <= today_date]
        df2 = df1[(df1['grass_date'] >= m1_sop) & (df1['grass_date'] <= today_date)]
        df2.drop_duplicates(inplace=True)
        df3 = df2.merge(color[['sku_id', 'inventory_value_usd', 'grass_date']],
                        on=['sku_id', 'grass_date'], how='left')
        df4 = df3.groupby(['category_cluster', 'supplier_name'])['inventory_value_usd'].\
            sum().reset_index()
        df4 = df4.sort_values(['supplier_name', 'inventory_value_usd'], ascending=False)
        df5 = df4.groupby('supplier_name')['supplier_name', 'category_cluster'].head(1)
        df5 = df5[['category_cluster', 'supplier_name']]

        latest_cogs = get_latest_cogs(cogs)

        replenishment_to_date = replenishment[
            (replenishment['grass_date'] >= m6_sop) & (replenishment['grass_date'] <= today_date)]

        m6_color = m6_df.merge(color, on=['sku_id', 'grass_date'], how='left')
        m6_color0 = m6_color.groupby(['supplier_name', 'grass_date'])['inventory_value_usd'].\
            sum().reset_index()
        m6_color1 = df5.merge(m6_color0, on='supplier_name', how='left')
        m6_color1['grass_date'] = m6_color1['grass_date'].dt.strftime('%Y-%m-%d')
        m6_color_pivot = pd.pivot_table(m6_color1, values='inventory_value_usd',
                                        index=['category_cluster', 'supplier_name'],
                                        columns='grass_date').reset_index()

        daily_cogs = replenishment_to_date.merge(latest_cogs[['sku_id', 'stored_value_fifo_per_qty']],
                                                 on='sku_id')
        daily_cogs['daily_cogs'] = daily_cogs['stored_value_fifo_per_qty'] * daily_cogs['total_sales']
        daily_cogs0 = m6_df.merge(daily_cogs[['sku_id', 'grass_date', 'daily_cogs']],
                                  on=['sku_id', 'grass_date'], how='left')
        daily_cogs1 = daily_cogs0.groupby(['supplier_name', 'grass_date'])['daily_cogs'].sum().reset_index()
        daily_cogs1['daily_cogs_usd'] = daily_cogs1['daily_cogs'] / country_x_rate
        daily_cogs2 = df5.merge(daily_cogs1, on=['supplier_name'], how='left')
        daily_cogs2 = daily_cogs2.groupby(by=['category_cluster', 'supplier_name',
                                              'grass_date'])['daily_cogs_usd'].sum().reset_index()
        daily_cogs2['grass_date'] = daily_cogs2['grass_date'].dt.strftime('%Y-%m-%d')
        daily_cogs_pivot = pd.pivot_table(daily_cogs2, values='daily_cogs_usd',
                                          index=['category_cluster', 'supplier_name'],
                                          columns='grass_date').reset_index()

        purchases['pr_created_date'] = pd.to_datetime(purchases['pr_created_date'])
        purchases0 = purchases.groupby(['sku_id', 'pr_created_date']).sum().reset_index()
        purchases1 = m6_df.merge(purchases0, left_on=['sku_id', 'grass_date'],
                                 right_on=['sku_id', 'pr_created_date'], how='left')
        purchases1['po_value'] = purchases1['po_value'].apply(pd.to_numeric, errors='coerce')
        # po_value is decimal, country_x_rate is float
        purchases1['po_value_usd'] = purchases1['po_value'] / country_x_rate
        purchases2 = purchases1.groupby(['supplier_name', 'grass_date'])['po_value_usd'].\
            sum().reset_index()
        purchases3 = df5.merge(purchases2, on='supplier_name', how='left')
        purchases3 = purchases3.groupby(by=['category_cluster', 'supplier_name',
                                            'grass_date'])['po_value_usd'].sum().reset_index()
        purchases3['grass_date'] = purchases3['grass_date'].dt.strftime('%Y-%m-%d')
        purchases_pivot = pd.pivot_table(purchases3, values='po_value_usd',
                                         index=['category_cluster', 'supplier_name'],
                                         columns='grass_date').reset_index()

        accounts_payable_df = m6_df1[m6_df1['grass_date'] <= today_date]
        accounts_payable_df.drop_duplicates(inplace=True)
        pr_po['inbound_time'] = pd.to_datetime(pr_po['inbound_time'])
        ap_inbounds = pr_po[pr_po['inbound_time'] <= today_date]
        ap_inbounds0 = ap_inbounds.merge(latest_cogs[['sku_id', 'stored_value_fifo_per_qty']],
                                         how='left', on='sku_id')
        ap_inbounds0 = ap_inbounds0.merge(accounts_payable_df[['sku_id', 'grass_date', 'payment_days']],
                                          right_on=['sku_id', 'grass_date'],
                                          left_on=['sku_id', 'inbound_time'],
                                          how='left')
        ap_inbounds0['payment_days'] = ap_inbounds0['payment_days'].fillna(0)
        ap_inbounds0['forward_days'] = pd.to_timedelta(ap_inbounds0['payment_days'], 'd')
        ap_inbounds0['ib_date_range'] = ap_inbounds0['inbound_time'] + ap_inbounds0['forward_days']
        ap_inbounds0['payment_pending'] = ap_inbounds0['stored_value_fifo_per_qty'] * ap_inbounds0['inbound_quantity']
        ap_inbounds0['payment_paid'] = -ap_inbounds0['stored_value_fifo_per_qty'] * ap_inbounds0['inbound_quantity']

        beg_payable['grass_date'] = m6_sop
        beg_payable.rename(columns={'end_payables': 'temp_payables'}, inplace=True)
        beg_payable.fillna(0, inplace=True)
        beg_payable['grass_date'] = pd.to_datetime(beg_payable['grass_date'])
        payment_pending_df = ap_inbounds0[['sku_id', 'grass_date', 'payment_pending']]
        payment_paid_df = ap_inbounds0[['sku_id', 'ib_date_range', 'payment_paid']]

        payment_pending_df.columns = ['sku_id', 'grass_date', 'payment_pending']
        payment_paid_df.columns = ['sku_id', 'grass_date', 'payment_paid']
        payment_pending_df = payment_pending_df.groupby(['sku_id', 'grass_date']).sum().reset_index()
        payment_paid_df = payment_paid_df.groupby(['sku_id', 'grass_date']).sum().reset_index()
        accounts_payable_df0 = accounts_payable_df.merge(beg_payable,
                                                         on=['sku_id', 'grass_date'], how='left')
        accounts_payable_df1 = accounts_payable_df0.merge(payment_pending_df,
                                                          on=['sku_id', 'grass_date'], how='left')
        accounts_payable_df2 = accounts_payable_df1.merge(payment_paid_df,
                                                          on=['sku_id', 'grass_date'], how='left')
        accounts_payable_df3 = accounts_payable_df2.sort_values('grass_date')
        accounts_payable_df3.fillna(0, inplace=True)
        accounts_payable_df3['total_ap'] = accounts_payable_df3['temp_payables'] \
                                           + accounts_payable_df3['payment_pending'] \
                                           + accounts_payable_df3['payment_paid']
        accounts_payable_df4 = accounts_payable_df3.copy()
        accounts_payable_df4 = accounts_payable_df4.groupby(['sku_id', 'grass_date'])[
            'temp_payables', 'payment_pending', 'payment_paid', 'total_ap'].sum()

        accounts_payable_df5 = accounts_payable_df4.reset_index()
        accounts_payable_df5['total_ap_cumsum'] = accounts_payable_df5.groupby(['sku_id'])['total_ap'].cumsum()
        accounts_payable_df5['temp_cumsum'] = accounts_payable_df5 \
            .groupby(['sku_id'])['temp_payables'].cumsum()
        accounts_payable_df6 = accounts_payable_df5.groupby('sku_id').nth(0)['temp_payables'].reset_index()
        accounts_payable_df6.rename(columns={"temp_payables": "beg_temp"}, inplace=True)
        accounts_payable_df7 = accounts_payable_df6.merge(accounts_payable_df5, on=['sku_id'])
        accounts_payable_df7['total_ap'] = accounts_payable_df7['total_ap_cumsum'] \
                                           - accounts_payable_df7['temp_cumsum'] \
                                           + accounts_payable_df7['beg_temp']
        accounts_payable_df7['current_acct_payable_usd'] = accounts_payable_df7['total_ap'] / country_x_rate
        accounts_payable_df7.drop(['temp_cumsum', 'total_ap_cumsum', 'beg_temp'], axis=1, inplace=True)

        m6_payables = m6_df.merge(accounts_payable_df7[['sku_id', 'grass_date', 'current_acct_payable_usd']],
                                  on=['sku_id', 'grass_date'], how='left')
        m6_payables0 = m6_payables.groupby(['supplier_name', 'grass_date'])[
            'current_acct_payable_usd'].sum().reset_index()
        m6_payables1 = df5.merge(m6_payables0, on='supplier_name', how='left')
        m6_payables1 = m6_payables1[m6_payables1['grass_date'] >= m6_sop]
        m6_payables1['grass_date'] = m6_payables1['grass_date'].dt.strftime('%Y-%m-%d')
        m6_payables_pivot = pd.pivot_table(m6_payables1, values='current_acct_payable_usd',
                                           columns='grass_date',
                                           index=['category_cluster', 'supplier_name']).reset_index()

        latest_date = df2['grass_date'].sort_values(ascending=False).head(1).values[0]
        initial_tab = df2[df2['grass_date'] == latest_date]
        latest_soh = color[color['grass_date'] == latest_date]
        initial_tab0 = initial_tab.merge(latest_soh[['sku_id', 'stock_on_hand', 'grass_date']],
                                         on=['sku_id', 'grass_date'], how='left')
        initial_tab1 = initial_tab0[initial_tab0['stock_on_hand'] > 0]
        tracking_tab = initial_tab1.groupby(['supplier_name'])['sku_id'].count().reset_index()
        tracking_tab.rename(columns={'sku_id': 'whs_sku_count'}, inplace=True)
        tracking_tab = df5.merge(tracking_tab, on='supplier_name', how='left')
        tracking_tab = tracking_tab[['category_cluster', 'supplier_name', 'whs_sku_count']]

        color_inventory_count0 = initial_tab0.merge(color[['sku_id', 'grass_date', 'inventory_value_usd']],
                                                    on=['sku_id', 'grass_date'], how='left')
        color_inventory0 = color_inventory_count0.groupby('supplier_name')['stock_on_hand', 'inventory_value_usd']. \
            sum().reset_index()
        color_inventory0.rename(columns={'stock_on_hand': 'inventory_count'}, inplace=True)
        tracking_tab0 = tracking_tab.merge(color_inventory0, on=['supplier_name'], how='left')
        brands_df = df2[['supplier_name', 'brand']]
        brands_df = get_clean_brands(brands_df)
        brands_df = brands_df[brands_df.brand != 'nan']
        number_of_top_brands = 3
        supplier_names = set(brands_df['supplier_name'].to_list())
        brand_count_df = brands_df.groupby(['supplier_name', 'brand']).size()
        # get nlargest per subgroup
        brand_count_top_3_df = brand_count_df.groupby(['supplier_name'], group_keys=False) \
            .apply(lambda subgroup: subgroup.nlargest(3))
        # drop count column and reset_index
        brand_count_top_3_df = brand_count_top_3_df.reset_index().drop(columns=0)
        # final brands dataframe in desired output
        top_brands_cols = ['supplier_name', 'brand_1', 'brand_2', 'brand_3']
        top_brands_df = pd.DataFrame(columns=top_brands_cols)
        for supplier in supplier_names:
            slice_df = brand_count_top_3_df[(brand_count_top_3_df.supplier_name == supplier)].reset_index()
            three_brands = dict()
            for i in range(number_of_top_brands):
                key = 'brand_{}'.format(i + 1)
                try:
                    brand = slice_df.iloc[i].brand
                    three_brands[key] = brand
                except IndexError:
                    three_brands[key] = 'n.a.'
            row_list = [supplier, three_brands['brand_1'], three_brands['brand_2'],
                        three_brands['brand_3']]
            row_to_append = pd.DataFrame([row_list], columns=top_brands_cols)
            top_brands_df = top_brands_df.append(row_to_append)

        top_brands_df = top_brands_df.reset_index(drop=True)
        top_brands_df.fillna('n.a.', inplace=True)
        tracking_tab2 = tracking_tab0.merge(top_brands_df, on=['supplier_name'], how='left')

        path = '/ldap_home/yutong.zou/{c1}/weekly_WC_{c2}.xlsx'.format(c1=country, c2=country)
        wb_obj = openpyxl.load_workbook(path)
        work_sheet = wb_obj[country]
        m_row = work_sheet.max_row
        supplier_highlight = []
        for i in range(4, m_row + 1):
            cell_obj = work_sheet.cell(row=i, column=1)
            if cell_obj.value in (None, 'TOTAL'):
                continue
            supplier_highlight.append(cell_obj.value)

        latest_payment_days = df2.sort_values(['grass_date'], ascending=False). \
            groupby(['supplier_name'])['payment_days'].first()
        tracking_df0 = tracking_tab2.merge(latest_payment_days, on=['supplier_name'], how='left')
        tracking_df1 = tracking_df0[tracking_df0['supplier_name'].isin(supplier_highlight)]

        main_ws = wb_obj['Tracking']
        main_ws['B1'] = today_date
        for df_index, row in tracking_df1.iterrows():
            for col_index in range(tracking_df1.shape[1]):
                cell_index = get_column_letter(col_index + 1) + str(df_index + 4)
                main_ws[cell_index] = row[col_index]

        write_dict = {
            'Daily Accounts Payable': m6_payables_pivot,
            'Daily Inventory Value': m6_color_pivot,
            'Daily COGS': daily_cogs_pivot,
            'Daily Inbounds': purchases_pivot
        }

        supplier_info_cols = {
            "category_cluster": "A",
            "supplier_name": "B",
        }

        start_date, days_in_months = get_days_in_months(m6_payables_pivot)
        append_to_excel(write_dict, supplier_info_cols, start_date, days_in_months, wb_obj)

        wb_obj.save(country_folder_path + '/week_{}_sample.xlsx'.format(country))

        # color_inv_pivot = spark.createDataFrame(m6_color_pivot)
        # acct_payable_pivot = spark.createDataFrame(m6_payables_pivot)
        # daily_cogs_pivot = spark.createDataFrame(daily_cogs_pivot)
        # purchases_pivot = spark.createDataFrame(purchases_pivot)
        # tracking_df1 = spark.createDataFrame(tracking_df1)

        m6_color_pivot.to_excel(country_folder_path + '/inventory_value_pivot.xlsx')
        m6_payables_pivot.to_excel(country_folder_path + '/acct_payable_pivot.xlsx')
        daily_cogs_pivot.to_excel(country_folder_path + '/cogs_pivot.xlsx')
        purchases_pivot.to_excel(country_folder_path + '/purchases_pivot.xlsx')
        tracking_df0.to_excel(country_folder_path + '/tracking_tab.xlsx')


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('-c', '--countries', nargs='+', default=['ID', 'MY', 'TH', 'VN', 'TW', 'PH'])
    parser.add_argument("-d", "--date",
                        default=dt.date.today(),
                        type=lambda d: dt.datetime.strptime(d, '%Y%m%d').date(),
                        help="Date in the format yyyymmdd")
    parser.add_argument('-q', '--queried', help='has data been queried?', default=False)
    args = parser.parse_args()
    main(args.date, args.countries, args.queried)
