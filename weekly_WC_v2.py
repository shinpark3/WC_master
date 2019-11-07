'''
This script generates a report to track the performance of each supplier,
and their daily inventory, payable, COGS and inbounds based on an excel template
'''

import os
import argparse
import calendar
import subprocess
import sys
import datetime as dt
import numpy as np
import pandas as pd
import dateutil.relativedelta
import openpyxl
import yaml
from openpyxl.utils import get_column_letter
from pyspark.sql import SparkSession
import time

start_time = time.time()

reload(sys)
sys.setdefaultencoding('utf8')

SPARK = SparkSession \
    .builder \
    .enableHiveSupport() \
    .getOrCreate()
SPARK.sql('use shopee')


def get_sop(current_month):
    '''
    This function returns the SOP (Start of Period) for the
    previous month which is denoted as the first day of the last month (SOP)
    :return SOP: start of period
    '''
    two_months_back = current_month - dateutil.relativedelta.relativedelta(months=2)
    m2_sop = two_months_back.replace(day=1)
    return m2_sop


def get_eoms(today_date):
    m0_eom = today_date
    m1_eom = today_date.replace(day=1) - dateutil.relativedelta.relativedelta(days=1)
    m2_eom = m1_eom.replace(day=1) - dateutil.relativedelta.relativedelta(days=1)
    return [m0_eom, m1_eom, m2_eom]


def get_days_in_months(today_date):
    '''
    Returns the number of days in each month assuming there are only three months
    The last number will always be 30 if today is not the end of the month
    :return: days_in_month (a list of number of days in each month)
    '''
    start_date = get_sop(today_date)
    days_in_months = [0,
                      pd.Period(start_date.strftime("%Y-%m-%d")).days_in_month,
                      pd.Period((start_date + pd.offsets.MonthBegin(1)).
                                strftime("%Y-%m-%d")).days_in_month]
    if today_date.day != calendar.monthrange(today_date.year, today_date.month)[1]:
        days_in_months.append(30)
    else:
        days_in_months.append(today_date.day)
    return days_in_months


def append_to_excel(write_dict, supplier_info_cols, today_date, days_in_months, wb_obj):
    '''
    This function paste the values in specified tables into the template,
    and automatically computes the averages and sums of one month
    :param write_dict: a dictionary of pivot tables
    :param supplier_info_cols: the columns of supplier category and names
    :param today_date: today date
    :param days_in_months: number of days in each month
    :param wb_obj: excel template to be written
    :return:
    '''
    start_row = 3
    start_date = get_sop(today_date)
    for sheet_name, pivot_df in write_dict.items():
        work_sheet = wb_obj[sheet_name]
        work_sheet['F2'] = start_date
        for df_index, row in pivot_df.iterrows():
            row_index = df_index + start_row
            for df_col_name, excel_letter in supplier_info_cols.items():
                cell_index = excel_letter + str(row_index)
                work_sheet[cell_index] = row[df_col_name]
            for column in range(6, pivot_df.shape[1] + 4):
                day_index = column - 4
                column_letter = get_column_letter(column)
                cell_index = column_letter + str(row_index)
                work_sheet[cell_index] = row[day_index]
            for column in range(3, 6):
                column_letter = get_column_letter(column)
                month_index = column - 2
                cell_index = column_letter + str(row_index)
                if month_index < len(days_in_months) - 1:
                    start_col = 2 + sum(days_in_months[:month_index])
                    end_col = 2 + sum(days_in_months[:month_index + 1])
                    if sheet_name in ['Daily Accounts Payable', 'Daily Inventory Value']:
                        work_sheet[cell_index] = row[start_col:end_col].mean()
                    else:
                        work_sheet[cell_index] = row[start_col:end_col].sum()
                else:
                    end_col = len(row)
                    start_col = end_col - days_in_months[month_index]
                    if sheet_name in ['Daily Accounts Payable', 'Daily Inventory Value']:
                        work_sheet[cell_index] = row[start_col:end_col].mean()
                    else:
                        work_sheet[cell_index] = row[start_col:end_col].sum()
            for column in range(pivot_df.shape[1] + 4, work_sheet.max_column):
                column_letter = get_column_letter(column)
                cell_index = column_letter + str(row_index)
                work_sheet[cell_index] = None


def read_suppliers(filename):
    '''
    Read supplier list from yaml filename
    :return: dictionary of country, suppliers pair
    '''
    with open(filename, 'r') as stream:
        dictionary = yaml.load(stream, Loader=yaml.FullLoader)
    supplier_dict = {}
    for key, value in dictionary.items():
        supplier_dict[str(key)] = value
    return supplier_dict


def get_main_df(country, today_date, data_queried):
    '''
    Get the data for the last three months for specific country
    '''
    country_folder_path = './' + country + '/'
    if not os.path.exists(country_folder_path):
        os.makedirs(country_folder_path)
    input_file_path = "{}input_files".format(country_folder_path)
    if not os.path.exists(input_file_path):
        os.makedirs(input_file_path)
    if not data_queried:
        m2_sop = get_sop(today_date).strftime("%Y-%m-%d")
        today_date = today_date.strftime("%Y-%m-%d")
        query = '''
            SELECT category_cluster, supplier_name, sku_id, cdate, color,
                    cogs_usd, stock_on_hand, inbound_value_usd, acct_payables_usd, 
                    inventory_value_usd, payment_terms, brand, grass_date 
            FROM shopee_bi_sbs_mart
            WHERE
                supplier_name is NOT NULL
            AND
                grass_date >= date('{start_date}')
            AND
                grass_date <= date('{end_date}')
            AND
                purchase_type = 'Outright'
            AND 
                grass_region = '{cntry}'
            '''.format(start_date=m2_sop, end_date=today_date, cntry=country)
        print(query)
        main_df = SPARK.sql(query)
        write_to_csv(main_df, 'main_df', input_file_path + '/main_df.csv')
        main_df = main_df.toPandas()
    else:
        main_df = pd.read_csv(input_file_path + '/main_df.csv',
                              encoding='utf-8-sig', index_col=False)
    return main_df


def write_to_csv(query_df, report_directory_name, output_file_name):
    '''
    This function writes the spark dataframe to a csv file
    :param query_df: spark dataframe
    :type query_df: spark dataframe
    :param output_file_name: Name of Output File
    :type output_file_name: str
    :param report_directory_name: Name of Report Directory
    :type report_directory_name: str
    :return:
    '''
    print("Output File Name: ", output_file_name)
    req_headers = query_df.columns
    username = os.getcwd().split('/')[-1]
    output_filename = 'user/{u}/{d}/tmp.csv'.format(u=username, d=report_directory_name)
    local_filename = output_file_name
    query_df.write.format('csv').mode('overwrite'). \
        options(header='false', escape='"', encoding="UTF-8").save(output_filename)
    subprocess.call('/usr/share/hadoop/bin/hadoop fs -getmerge %s %s' % (
        output_filename, local_filename + "_tmp"), shell=True)
    subprocess.call('/usr/share/hadoop/bin/hadoop fs -rm -r %s' % output_filename, shell=True)
    db_input_file = open(local_filename + "_tmp", 'rb')
    print(db_input_file)
    db_output_file = open(local_filename, 'wb')
    db_output_file.write(','.join(req_headers) + '\n')
    for line in db_input_file:
        db_output_file.write(line)
    db_output_file.close()
    os.remove(local_filename + "_tmp")


def main(country, today_date, main_df, supplier_dict):
    '''
    Pull data; write to the template and save tracking table and
    pivot tables for inventory, COGS, inbounds and COGS
    :param country: country
    :param today_date: last day of the report
    :param main_df: main data frame
    :param supplier_dict: dictionary of country, suppliers list pair
    '''
    half_time = time.time()
    print("--- %s seconds --- generating data" % (half_time - start_time))
    country_folder_path = './' + country + '/'
    main_df['grass_date'] = pd.to_datetime(main_df['grass_date'])
    main_df['cdate'] = pd.to_datetime(main_df['cdate'])
    df_info = main_df[main_df['grass_date'] == today_date]

    df_inv_count = df_info.groupby('supplier_name')['stock_on_hand'].sum().reset_index()
    df_inv_count.columns = ['supplier_name', 'inv_count']

    df_inv_sorting = main_df.sort_values(['grass_date'], ascending=False) \
        .groupby(['supplier_name'])['supplier_name', 'grass_date'] \
        .head(1).dropna(subset=['supplier_name'])
    df_inv_sorting0 = df_inv_sorting.merge(main_df, on=['supplier_name', 'grass_date'])
    df_inv_sorting1 = df_inv_sorting0.groupby(['category_cluster', 'supplier_name'])[
        'inventory_value_usd'].sum().reset_index()
    df_inv_sorting2 = df_inv_sorting1.sort_values(['supplier_name', 'inventory_value_usd'],
                                                  ascending=False).reset_index() \
        .groupby('supplier_name')['supplier_name', 'category_cluster'].head(1).reset_index()
    df_inv_sorting2 = df_inv_sorting2[['category_cluster', 'supplier_name']]

    df_inv_count = df_inv_sorting2.merge(df_inv_count, on=['supplier_name'], how='left')
    df_inv_sum = df_info.groupby('supplier_name')['inventory_value_usd'].sum().reset_index()

    df_sku_count = df_info.groupby('supplier_name')['sku_id'].count().reset_index()
    df_sku_count.columns = ['supplier_name', 'no_skus_WH']
    df_info0 = df_sku_count.merge(df_inv_count, on=['supplier_name'], how='left')

    df_payment = df_info.groupby('supplier_name')['supplier_name', 'payment_terms']. \
        head(1).reset_index(drop=True)
    df_last_month = main_df[(main_df['grass_date'] >= today_date - dt.timedelta(days=30))
                            & (main_df['grass_date'] <= today_date)]
    df_last_month.drop_duplicates(inplace=True)
    brands_df = df_last_month[['supplier_name', 'brand']]
    brands_df = brands_df[brands_df.brand != 'nan']
    number_of_top_brands = 3
    supplier_names = set(brands_df['supplier_name'].to_list())
    brand_count_df = brands_df.groupby(['supplier_name', 'brand']).size()
    # get nlargest per subgroup
    brand_count_top_3_df = brand_count_df.groupby(['supplier_name'], group_keys=False).apply(
        lambda subgroup: subgroup.nlargest(number_of_top_brands))
    # drop count column and reset_index
    brand_count_top_3_df = brand_count_top_3_df.reset_index().drop(columns=0)
    # final brands dataframe in desired output
    top_brands_cols = ['supplier_name', 'brand_1', 'brand_2', 'brand_3']
    top_brands_df = pd.DataFrame(columns=top_brands_cols)
    for supplier in supplier_names:
        slice_df = brand_count_top_3_df[(brand_count_top_3_df.supplier_name == supplier)]. \
            reset_index()
        three_brands = dict()
        for i in range(number_of_top_brands):
            key = 'brand_{}'.format(i + 1)
            try:
                brand = slice_df.iloc[i].brand
                three_brands[key] = brand
            except IndexError:
                three_brands[key] = 'n.a.'
        row_list = [supplier, three_brands['brand_1'], three_brands['brand_2'],
                    three_brands['brand_3']]
        row_to_append = pd.DataFrame([row_list], columns=top_brands_cols)
        top_brands_df = top_brands_df.append(row_to_append)

    top_brands_df = top_brands_df.reset_index(drop=True)
    top_brands_df.fillna('n.a.', inplace=True)
    df_info1 = df_info0.merge(df_inv_sum, on=['supplier_name'], how='left')
    df_info2 = df_info1.merge(top_brands_df, on=['supplier_name'], how='left')
    df_info3 = df_info2.merge(df_payment, on=['supplier_name'], how='left')

    df_repln = main_df[['supplier_name', 'inbound_value_usd', 'cdate', 'grass_date', 'color']]
    df_repln['is_replenishment'] = np.where(df_repln['cdate']
                                            < df_repln['grass_date'].astype('datetime64[M]'), 1, 0)
    df_repln['inb_repln_usd'] = df_repln['inbound_value_usd'] * df_repln['is_replenishment']
    df_repln['is_green_repln'] = np.where((df_repln['color'] == 'Green')
                                          & (df_repln['is_replenishment'] == 1), 1, 0)
    df_repln['green_repln_usd'] = df_repln['inbound_value_usd'] * df_repln['is_green_repln']
    df_repln1 = df_repln.groupby(['supplier_name',
                                  pd.Grouper(key='grass_date', freq='M')]) \
        [['inb_repln_usd', 'green_repln_usd']].sum()
    df_repln2 = df_repln1.pivot_table(values='inb_repln_usd', index='supplier_name', columns='grass_date') \
        .reset_index()
    df_green_repln = df_repln1.pivot_table(values='green_repln_usd', index='supplier_name', columns='grass_date') \
        .reset_index()
    df_repln2.columns = ['supplier_name', 'inb_repln_m2', 'inb_repln_m1', 'inb_repln_m0']
    df_repln2.drop(['inb_repln_m0'], axis=1, inplace=True)
    df_green_repln.columns = ['supplier_name', 'green_repln_m2', 'green_repln_m1', 'green_repln_m0']
    df_green_repln.drop(['green_repln_m0'], axis=1, inplace=True)

    # re-compute last 30d as m0
    df_repln1_m0 = df_repln[df_repln['grass_date'] >= today_date-dt.timedelta(days=29)]
    df_repln_m0 = df_repln1_m0.groupby(['supplier_name'])[['inb_repln_usd']].sum()
    df_repln3 = df_repln2.merge(df_repln_m0, on=['supplier_name'], how='left')
    df_repln3.columns = ['supplier_name', 'inb_repln_m2', 'inb_repln_m1', 'inb_repln_m0']

    df_green_repln_m0 = df_repln1_m0.groupby(['supplier_name'])[['green_repln_usd']].sum()
    df_green_repln1 = df_green_repln.merge(df_green_repln_m0, on=['supplier_name'], how='left')
    df_green_repln1.columns = ['supplier_name', 'green_repln_m2', 'green_repln_m1', 'green_repln_m0']

    eoms = get_eoms(today_date)
    df_eom_inv = main_df[['supplier_name', 'inventory_value_usd', 'grass_date', 'color']]
    df_eom_inv0 = df_eom_inv[(df_eom_inv['grass_date']).isin(eoms)]
    df_eom_inv0['is_black_inv'] = np.where(df_eom_inv0['color'] == 'Black', 1, 0)
    df_eom_inv0['black_inv_usd'] = df_eom_inv0['is_black_inv'] * df_eom_inv0['inventory_value_usd']

    df_black_inv0 = df_eom_inv0.groupby(['supplier_name', 'grass_date']) \
        [['black_inv_usd']].sum()\
        .pivot_table(values='black_inv_usd', index='supplier_name', columns='grass_date')\
        .reset_index()
    if df_black_inv0.shape[1] == 3:
        print('Warning: There is likely to be no data for the date you specified')
    df_black_inv0.columns = ['supplier_name', 'black_inv_m2', 'black_inv_m1', 'black_inv_m0']
    df_eom_inv1 = df_eom_inv0.groupby(['supplier_name', 'grass_date']) \
        [['inventory_value_usd']].sum()\
        .pivot_table(values='inventory_value_usd', index='supplier_name', columns='grass_date')\
        .reset_index()
    df_eom_inv1.columns = ['supplier_name', 'eom_inv_m2', 'eom_inv_m1', 'eom_inv_m0']
    df_eom_inv2 = df_eom_inv1.merge(df_black_inv0, on=['supplier_name'], how='left')

    df_total_sum = main_df.groupby(['supplier_name', 'grass_date']).sum().reset_index()
    df_monthly_inbound = df_total_sum.groupby(['supplier_name',
                                               pd.Grouper(key='grass_date', freq='M')]) \
        [['inbound_value_usd']].sum() \
        .pivot_table(values='inbound_value_usd', index='supplier_name', columns='grass_date') \
        .reset_index()
    df_monthly_inbound.columns = ['supplier_name', 'inb_m2', 'inb_m1', 'inb_m0']
    df_monthly_inbound.drop(['inb_m0'], axis=1, inplace=True)

    df_total_sum_m0 = df_total_sum[df_total_sum['grass_date'] >= today_date-dt.timedelta(days=29)]
    df_monthly_inbound_m0 = df_total_sum_m0.groupby(['supplier_name'])[['inbound_value_usd']].sum()
    df_monthly_inbound = df_monthly_inbound.merge(df_monthly_inbound_m0, on=['supplier_name'], how='left')
    df_monthly_inbound.columns = ['supplier_name', 'inb_m2', 'inb_m1', 'inb_m0']

    df_total_sum['grass_date'] = df_total_sum['grass_date'].dt.strftime('%Y-%m-%d')
    df_cogs = pd.pivot_table(df_total_sum, values='cogs_usd', index='supplier_name',
                             columns='grass_date').reset_index()
    df_cogs = df_inv_sorting2.merge(df_cogs, on=['supplier_name'], how='right')
    df_payable = pd.pivot_table(df_total_sum, values='acct_payables_usd',
                                index='supplier_name', columns='grass_date').reset_index()
    df_payable = df_inv_sorting2.merge(df_payable, on=['supplier_name'], how='right')
    df_inv_value = pd.pivot_table(df_total_sum, values='inventory_value_usd',
                                  index='supplier_name', columns='grass_date').reset_index()
    df_inv_value = df_inv_sorting2.merge(df_inv_value, on=['supplier_name'], how='right')
    df_inbound = pd.pivot_table(df_total_sum, values='inbound_value_usd',
                                index='supplier_name', columns='grass_date').reset_index()
    df_inbound = df_inv_sorting2.merge(df_inbound, on=['supplier_name'], how='right')

    df_info3 = df_info3[['category_cluster', 'supplier_name', 'no_skus_WH',
                         'inv_count', 'inventory_value_usd', 'brand_1',
                         'brand_2', 'brand_3', 'payment_terms']]

    if supplier_dict.get(country) is None:
        df_info4 = df_info3
    else:
        supplier_highlight = supplier_dict.get(country)
        df_info4 = df_info3[df_info3['supplier_name'].isin(supplier_highlight)] \
            .reset_index(drop=True)

    df_info5 = df_info4.merge(df_monthly_inbound, on=['supplier_name'], how='left')\
        .merge(df_repln3, on=['supplier_name'], how='left')\
        .merge(df_green_repln1, on=['supplier_name'], how='left')\
        .merge(df_eom_inv2, on=['supplier_name'], how='left')

    df_info5['inb_repln_m2_perc'] = df_info5['inb_repln_m2']/ df_info5['inb_m2']
    df_info5['inb_repln_m1_perc'] = df_info5['inb_repln_m1'] / df_info5['inb_m1']
    df_info5['inb_repln_m0_perc'] = df_info5['inb_repln_m0'] / df_info5['inb_m0']
    df_info5['green_repln_m2_perc'] = df_info5['green_repln_m2'] / df_info5['inb_repln_m2']
    df_info5['green_repln_m1_perc'] = df_info5['green_repln_m1'] / df_info5['inb_repln_m1']
    df_info5['green_repln_m0_perc'] = df_info5['green_repln_m0'] / df_info5['inb_repln_m0']
    df_info5['black_inv_m2_perc'] = df_info5['black_inv_m2'] / df_info5['eom_inv_m2']
    df_info5['black_inv_m1_perc'] = df_info5['black_inv_m1'] / df_info5['eom_inv_m1']
    df_info5['black_inv_m0_perc'] = df_info5['black_inv_m0'] / df_info5['eom_inv_m0']
    df_info5.fillna('N/A', inplace=True)
    df_info5 = df_info5[['category_cluster', 'supplier_name', 'no_skus_WH', 'inv_count',
                         'inventory_value_usd', 'brand_1', 'brand_2', 'brand_3', 'payment_terms',
                         'inb_repln_m2_perc', 'inb_repln_m1_perc', 'inb_repln_m0_perc',
                         'green_repln_m2_perc', 'green_repln_m1_perc', 'green_repln_m0_perc',
                         'black_inv_m2', 'black_inv_m1', 'black_inv_m0',
                         'black_inv_m2_perc', 'black_inv_m1_perc', 'black_inv_m0_perc']]

    path = country_folder_path + 'Weekly_wc_template.xlsx'
    wb_obj = openpyxl.load_workbook(path)
    main_ws = wb_obj['Tracking']
    main_ws['B1'] = today_date
    for df_index, row in df_info5.iterrows():
        for col_index in range(9):
            cell_index = get_column_letter(col_index + 1) + str(df_index + 4)
            main_ws[cell_index] = row[col_index]
        for col_index in range(9,21):
            cell_index = get_column_letter(col_index + 6 + 1) + str(df_index + 4)
            main_ws[cell_index] = row[col_index]

    write_dict = {
        'Daily Accounts Payable': df_payable,
        'Daily Inventory Value': df_inv_value,
        'Daily COGS': df_cogs,
        'Daily Inbounds': df_inbound
    }
    supplier_info_cols = {
        "category_cluster": "A",
        "supplier_name": "B",
    }
    days_in_months = get_days_in_months(today_date)
    append_to_excel(write_dict, supplier_info_cols, today_date, days_in_months, wb_obj)
    wb_obj.save(country_folder_path + '/month_{}_sample_v2.xlsx'.format(country))

    df_info3.to_csv(country_folder_path + '{}_tracking_tab_v2.csv'.format(country),
                    encoding="utf-8-sig")
    df_payable.to_csv(country_folder_path + '{}_payables_v2.csv'.format(country),
                      encoding="utf-8-sig")
    df_cogs.to_csv(country_folder_path + '{}_cogs_v2.csv'.format(country),
                   encoding="utf-8-sig")
    df_inv_value.to_csv(country_folder_path + '{}_inventory_value_v2.csv'.format(country),
                        encoding="utf-8-sig")
    df_inbound.to_csv(country_folder_path + '{}_inbound_v2.csv'.format(country),
                      encoding="utf-8-sig")
    result_dict = {'tracking': df_info3, 'payable': df_payable, 'cogs': df_cogs,
                   'inventory': df_inv_value, 'inbound': df_inbound, 'tracking_yaml': df_info4}
    final_time = time.time()
    print("--- %s seconds ---processing time" % (final_time - half_time))
    return result_dict


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('-c', '--countries', nargs='+', default=['ID', 'MY', 'TH', 'VN', 'TW', 'PH'])
    parser.add_argument("-d", "--date",
                        default=dt.date.today(),
                        type=lambda d: dt.datetime.strptime(d, '%Y%m%d').date(),
                        help="Date in the format yyyymmdd")
    parser.add_argument('-q', '--queried', help='has data been queried?', default=False)
    args = parser.parse_args()
    supplier_dict0 = read_suppliers('./suppliers.yaml')
    for country0 in args.countries:
        main_df0 = get_main_df(country0, args.date, args.queried)
        main(country0, args.date, main_df0, supplier_dict0)
