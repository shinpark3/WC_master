'''
This script generates a report to track the performance of each supplier,
and their daily inventory, payable, COGS and inbounds based on an excel template
'''

import os
import argparse
import calendar
import subprocess
import sys
import datetime as dt
import pandas as pd
import dateutil.relativedelta
import yaml
import openpyxl
from openpyxl.utils import get_column_letter
from pyspark.sql import SparkSession
from pyspark.sql.functions import to_date
import pyspark.sql.functions as F
from pyspark.sql.window import Window
from pyspark.sql.functions import col
from pyspark import SparkContext
from pyspark.sql import SQLContext
from pyspark.sql.types import StructType, StructField
from pyspark.sql.types import StringType, FloatType, IntegerType
import time
start_time = time.time()

sc = SparkContext.getOrCreate()
sqlContext = SQLContext(sc)

reload(sys)
sys.setdefaultencoding('utf8')

spark = SparkSession \
    .builder \
    .enableHiveSupport() \
    .getOrCreate()
spark.sql('use shopee')
# spark.sql('set spark.sql.shuffle.partitions=50')

def get_sop(current_month):
    '''
    This function returns the SOP (Start of Period) for the
    previous month which is denoted as the first day of the last month (SOP)
    :return SOP: start of period
    '''
    two_months_back = current_month - dateutil.relativedelta.relativedelta(months=2)
    m2_sop = two_months_back.replace(day=1)
    return m2_sop


def get_days_in_months(today_date):
    '''
    Returns the number of days in each month assuming there are only three months
    The last number will always be 30 if today is not the end of the month
    :return: days_in_month (a list of number of days in each month)
    '''
    start_date = get_sop(today_date)
    days_in_months = [0,
                      pd.Period(start_date.strftime("%Y-%m-%d")).days_in_month,
                      pd.Period((start_date + pd.offsets.MonthBegin(1)).
                                strftime("%Y-%m-%d")).days_in_month]
    if today_date.day != calendar.monthrange(today_date.year, today_date.month)[1]:
        days_in_months.append(30)
    else:
        days_in_months.append(today_date.day)
    return days_in_months


def append_to_excel(write_dict, supplier_info_cols, today_date, days_in_months, wb_obj):
    '''
    This function paste the values in specified tables into the template,
    and automatically computes the averages and sums of one month
    :param write_dict: a dictionary of pivot tables
    :param supplier_info_cols: the columns of supplier category and names
    :param today_date: today date
    :param days_in_months: number of days in each month
    :param wb_obj: excel template to be written
    :return:
    '''
    start_row = 3
    start_date = get_sop(today_date)
    for sheet_name, pivot_df in write_dict.items():
        work_sheet = wb_obj[sheet_name]
        work_sheet['F2'] = start_date
        for df_index, row in pivot_df.iterrows():
            row_index = df_index + start_row
            for df_col_name, excel_letter in supplier_info_cols.items():
                cell_index = excel_letter + str(row_index)
                work_sheet[cell_index] = row[df_col_name]
            for column in range(6, pivot_df.shape[1] + 4):
                day_index = column - 4
                column_letter = get_column_letter(column)
                cell_index = column_letter + str(row_index)
                work_sheet[cell_index] = row[day_index]
            for column in range(3, 6):
                column_letter = get_column_letter(column)
                month_index = column - 2
                cell_index = column_letter + str(row_index)
                if month_index < len(days_in_months) - 1:
                    start_col = 2 + sum(days_in_months[:month_index])
                    end_col = 2 + sum(days_in_months[:month_index + 1])
                    if sheet_name in ['Daily Accounts Payable', 'Daily Inventory Value']:
                        work_sheet[cell_index] = row[start_col:end_col].mean()
                    else:
                        work_sheet[cell_index] = row[start_col:end_col].sum()
                else:
                    end_col = len(row)
                    start_col = end_col - days_in_months[month_index]
                    if sheet_name in ['Daily Accounts Payable', 'Daily Inventory Value']:
                        work_sheet[cell_index] = row[start_col:end_col].mean()
                    else:
                        work_sheet[cell_index] = row[start_col:end_col].sum()
            for column in range(pivot_df.shape[1] + 4, work_sheet.max_column):
                column_letter = get_column_letter(column)
                cell_index = column_letter + str(row_index)
                work_sheet[cell_index] = None


def read_suppliers(filename):
    '''
    Read supplier list from yaml filename
    :return: dictionary of country, suppliers pair
    '''
    with open(filename, 'r') as stream:
        dictionary = yaml.load(stream, Loader=yaml.FullLoader)
    supplier_dict = {}
    for key, value in dictionary.items():
        supplier_dict[str(key)] = value

    return supplier_dict


def get_main_df(country, today_date, data_queried):
    '''
    Get the data for the last three months for specific country
    '''
    country_folder_path = './' + country + '/'
    if not os.path.exists(country_folder_path):
        os.makedirs(country_folder_path)

    input_file_path = "{}input_files".format(country_folder_path)
    if not os.path.exists(input_file_path):
        os.makedirs(input_file_path)

    if not data_queried:
        m2_sop = get_sop(today_date).strftime("%Y-%m-%d")
        today_date = today_date.strftime("%Y-%m-%d")
        query = '''
            SELECT category_cluster, supplier_name, sku_id, sourcing_status, 
                    cogs_usd, stock_on_hand, inbound_value_usd, acct_payables_usd, 
                    inventory_value_usd, payment_terms, brand, grass_date 
            from shopee_bi_sbs_mart
            where
                grass_date >= date('{start_date}')
            and
                grass_date <= date('{end_date}')
            AND
                purchase_type = 'Outright'
            and 
                grass_region = '{cntry}'
            '''.format(start_date=m2_sop, end_date=today_date, cntry=country)
        print(query)
        main_df = spark.sql(query)
        write_to_csv(main_df, input_file_path + '/main_df.csv', country)
    else:
        username = os.getcwd().split('/')[-1]
        my_schema = StructType([
            StructField("category_cluster", StringType()),
            StructField("supplier_name", StringType()),
            StructField("sku_id", StringType()),
            StructField("sourcing_status", StringType()),
            StructField("cogs_usd", FloatType()),
            StructField("stock_on_hand", IntegerType()),
            StructField("inbound_value_usd", FloatType()),
            StructField("acct_payables_usd", FloatType()),
            StructField("inventory_value_usd", FloatType()),
            StructField("payment_terms", StringType()),
            StructField("brand", StringType()),
            StructField("grass_date", StringType())
        ])
        input_file_path0 = 'hdfs://tl0/user/' + username + '/WC/' + country + '/input_files/main_df.csv'
        main_df = spark.read.csv(input_file_path0, header=True, schema=my_schema)
    return main_df


def write_to_csv(query_df, output_file_name, country):
    '''
    This function writes the spark dataframe to a csv file
    :param query_df: spark dataframe
    :type query_df: spark dataframe
    :param output_file_name: Name of Output File
    :type output_file_name: str
    :param country: country
    :type country: str
    :return:
    '''
    print("Output File Name: ", output_file_name)
    req_headers = query_df.columns
    username = os.getcwd().split('/')[-1]
    output_filename = 'user/{}/tmp.csv'.format(username)
    local_filename = output_file_name
    query_df.write.format('csv').mode('overwrite'). \
        options(header='false', escape='"', encoding="UTF-8").save(output_filename)
    output_file_path = 'hdfs://tl0/user/' + username + '/WC/' + country + '/input_files/main_df.csv'
    query_df.write.format('csv').mode('overwrite').save(output_file_path)
    subprocess.call('/usr/share/hadoop/bin/hadoop fs -getmerge %s %s' % (
        output_filename, local_filename + "_tmp"), shell=True)
    subprocess.call('/usr/share/hadoop/bin/hadoop fs -rm -r %s' % output_filename, shell=True)
    db_input_file = open(local_filename + "_tmp", 'rb')
    print(db_input_file)
    db_output_file = open(local_filename, 'wb')
    db_output_file.write(','.join(req_headers) + '\n')
    for line in db_input_file:
        db_output_file.write(line)
    db_output_file.close()
    os.remove(local_filename + "_tmp")


def main(country, today_date, main_df, supplier_dict):
    '''
    Pull data; write to the template and save tracking table and
    pivot tables for inventory, COGS, inbounds and COGS
    :param country: country
    :param today_date: last day of the report
    :param main_df: main data frame
    :param supplier_dict: a dictionary of country, suppliers pair
    '''
    half_time = time.time()
    print("--- %s seconds --- generating data" % (half_time - start_time))
    country_folder_path = './' + country + '/'
    main_df = main_df.withColumn('grass_date', to_date(col('grass_date')))
    df_info = main_df.filter(col('grass_date') == today_date)
    start_date = get_sop(today_date)
    dates = [(start_date + dt.timedelta(days=x)).strftime("%Y-%m-%d")
             for x in range((today_date - start_date).days + 1)]

    df_inv_count = df_info.groupBy('supplier_name')\
        .agg(F.sum('stock_on_hand').alias('inv_count'))

    df_inv_sorting = df_info.groupBy(['category_cluster', 'supplier_name'])\
        .agg(F.sum('inventory_value_usd').alias('inventory_value_usd'))\
        .sort(['inventory_value_usd'], ascending=False)\
        .groupBy('supplier_name')\
        .agg(F.first('category_cluster').alias('category_cluster'))\
        .select('category_cluster', 'supplier_name')

    df_inv_count = df_inv_sorting.join(df_inv_count, on=['supplier_name'], how='left')

    df_inv_sum = df_info.groupBy('supplier_name')\
        .agg(F.sum('inventory_value_usd').alias('inventory_value_usd'))

    df_sku_count = df_info.groupBy('supplier_name')\
        .agg(F.countDistinct('sku_id').alias('no_skus_WH'))

    df_info0 = df_sku_count.join(df_inv_count, on=['supplier_name'], how='left')

    df_payment = df_info.groupBy('supplier_name')\
        .agg(F.first('payment_terms').alias('payment_terms'))

    df_last_month = main_df.filter((col('grass_date') >= today_date - dt.timedelta(days=30))
                                   & (col('grass_date') <= today_date)).dropDuplicates()
    brands_df = df_last_month.select('supplier_name', 'brand')\
        .filter(col('brand') != 'nan')
    number_of_top_brands = 3
    supplier_names = set(brands_df.select('supplier_name')
                         .dropDuplicates()
                         .toPandas()['supplier_name']
                         .to_list())
    brand_count_df = brands_df.groupBy(['supplier_name', 'brand']).count()
    # get nlargest per subgroup
    window = Window.partitionBy(brand_count_df['supplier_name'])\
        .orderBy(brand_count_df['count'].desc())
    brand_count_top_3_df = brand_count_df.select('*', F.rank().over(window).alias('rank'))\
        .filter(col('rank') <= number_of_top_brands)\
        .drop('count', 'rank')
    brand_count_top_3_df = brand_count_top_3_df.toPandas()
    # final brands dataframe in desired output
    top_brands_cols = ['supplier_name', 'brand_1', 'brand_2', 'brand_3']
    top_brands_df = pd.DataFrame(columns=top_brands_cols)
    for supplier in supplier_names:
        slice_df = brand_count_top_3_df[(brand_count_top_3_df.supplier_name == supplier)].\
            reset_index()
        three_brands = dict()
        for i in range(number_of_top_brands):
            key = 'brand_{}'.format(i + 1)
            try:
                brand = slice_df.iloc[i].brand
                three_brands[key] = brand
            except IndexError:
                three_brands[key] = 'n.a.'
        row_list = [supplier, three_brands['brand_1'], three_brands['brand_2'],
                    three_brands['brand_3']]
        row_to_append = pd.DataFrame([row_list], columns=top_brands_cols)
        top_brands_df = top_brands_df.append(row_to_append)

    top_brands_df = top_brands_df.reset_index(drop=True)
    top_brands_df.fillna('n.a.', inplace=True)
    my_schema = StructType([StructField("supplier_name", StringType(), True),
                           StructField("brand_1", StringType(), True),
                           StructField("brand_2", StringType(), True),
                           StructField("brand_3", StringType(), True)])
    top_brands_df = sqlContext.createDataFrame(top_brands_df, schema=my_schema)

    df_info1 = df_info0.join(df_inv_sum, on=['supplier_name'], how='left')
    df_info2 = df_info1.join(top_brands_df, on=['supplier_name'], how='left')
    df_info3 = df_info2.join(df_payment, on=['supplier_name'], how='left')

    df_total_sum = main_df.groupBy(['supplier_name', 'grass_date'])\
        .agg(F.sum('cogs_usd').alias('cogs_usd'),
             F.sum('acct_payables_usd').alias('acct_payables_usd'),
             F.sum('inventory_value_usd').alias('inventory_value_usd'),
             F.sum('inbound_value_usd').alias('inbound_value_usd'))

    df_total_sum = df_total_sum.withColumn('grass_date',
                                           F.udf(lambda d: d.strftime('%Y-%m-%d'),
                                                 StringType())(col('grass_date')))

    col_names = ['category_cluster', 'supplier_name'] + dates
    df_cogs = df_total_sum.groupBy('supplier_name')\
        .pivot('grass_date')\
        .agg(F.first('cogs_usd'))\
        .join(df_inv_sorting, on=['supplier_name'], how='left')\
        .select(col_names)\
        .toPandas()

    df_payable = df_total_sum.groupBy('supplier_name')\
        .pivot('grass_date')\
        .agg(F.first('acct_payables_usd'))\
        .join(df_inv_sorting, on=['supplier_name'], how='left')\
        .select(col_names)\
        .toPandas()

    df_inv_value = df_total_sum.groupBy('supplier_name')\
        .pivot('grass_date')\
        .agg(F.first('inventory_value_usd'))\
        .join(df_inv_sorting, on=['supplier_name'], how='left')\
        .select(col_names)\
        .toPandas()

    df_inbound = df_total_sum.groupBy('supplier_name') \
        .pivot('grass_date') \
        .agg(F.first('inbound_value_usd'))\
        .join(df_inv_sorting, on=['supplier_name'], how='left')\
        .select(col_names)\
        .toPandas()

    df_info3 = df_info3.select('category_cluster', 'supplier_name', 'no_skus_WH',
                               'inv_count', 'inventory_value_usd', 'brand_1',
                               'brand_2', 'brand_3', 'payment_terms').toPandas()

    if supplier_dict.get(country) is None:
        df_info4 = df_info3
    else:
        supplier_highlight = supplier_dict.get(country)
        df_info4 = df_info3[df_info3['supplier_name'].isin(supplier_highlight)]\
            .reset_index(drop=True)

    path = country_folder_path + 'Weekly_wc_template.xlsx'
    wb_obj = openpyxl.load_workbook(path)
    main_ws = wb_obj['Tracking']
    main_ws['B1'] = today_date
    for df_index, row in df_info4.iterrows():
        for col_index in range(df_info4.shape[1]):
            cell_index = get_column_letter(col_index + 1) + str(df_index + 4)
            main_ws[cell_index] = row[col_index]

    write_dict = {
        'Daily Accounts Payable': df_payable,
        'Daily Inventory Value': df_inv_value,
        'Daily COGS': df_cogs,
        'Daily Inbounds': df_inbound
    }
    supplier_info_cols = {
        "category_cluster": "A",
        "supplier_name": "B",
    }
    days_in_months = get_days_in_months(today_date)
    append_to_excel(write_dict, supplier_info_cols, today_date, days_in_months, wb_obj)
    wb_obj.save(country_folder_path + '/month_{}_sample.xlsx'.format(country))

    df_info3.to_csv(country_folder_path + '{}_tracking_tab_v2.csv'.format(country),
                    encoding="utf-8-sig")
    df_payable.to_csv(country_folder_path + '{}_payables_v2.csv'.format(country),
                      encoding="utf-8-sig")
    df_cogs.to_csv(country_folder_path + '{}_cogs_v2.csv'.format(country),
                   encoding="utf-8-sig")
    df_inv_value.to_csv(country_folder_path + '{}_inventory_value_v2.csv'.format(country),
                        encoding="utf-8-sig")
    df_inbound.to_csv(country_folder_path + '{}_inbound_v2.csv'.format(country),
                      encoding="utf-8-sig")
    result_dict = {'tracking': df_info3, 'payable': df_payable, 'cogs': df_cogs,
                   'inventory': df_inv_value, 'inbound': df_inbound, 'tracking_yaml': df_info4}
    final_time = time.time()
    print("--- %s seconds ---processing time" % (final_time - half_time))
    return result_dict


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('-c', '--countries', nargs='+', default=['ID', 'MY', 'TH', 'VN', 'TW', 'PH'])
    parser.add_argument("-d", "--date",
                        default=dt.date.today(),
                        type=lambda d: dt.datetime.strptime(d, '%Y%m%d').date(),
                        help="Date in the format yyyymmdd")
    parser.add_argument('-q', '--queried', help='has data been queried?', default=False)
    args = parser.parse_args()
    supplier_dict0 = read_suppliers('./suppliers.yaml')
    for country0 in args.countries:
        main_df0 = get_main_df(country0, args.date, args.queried)
        main(country0, args.date, main_df0, supplier_dict0)
