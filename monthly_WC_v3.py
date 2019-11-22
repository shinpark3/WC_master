'''
This script generates a report to track the performance of each supplier,
and their daily inventory, payable, COGS and inbounds based on an excel template
v3: summary date ranges are of 90d, 90d, 90d, 30d
'''

import os
import argparse
import subprocess
import sys
import datetime as dt
import numpy as np
import pandas as pd
import dateutil.relativedelta
import openpyxl
import yaml
from openpyxl.utils import get_column_letter
from pyspark.sql import SparkSession
import time
import calendar

start_time = time.time()

reload(sys)
sys.setdefaultencoding('utf8')

SPARK = SparkSession \
    .builder \
    .enableHiveSupport() \
    .getOrCreate()
SPARK.sql('use shopee')


def get_sop(current_month):
    '''
    This function returns the SOP (Start of Period) for the
    previous month which is denoted as the first day of the last month (SOP)
    :return SOP: start of period
    '''
    m2_eop = get_eops(current_month)[0]
    m2_sop = m2_eop - dt.timedelta(days=89)
    return m2_sop


def get_eops(today_date):
    m0_eop = today_date
    m1_eop = today_date.replace(day=1) - dateutil.relativedelta.relativedelta(days=1)
    m2_eop = m1_eop.replace(day=1) - dateutil.relativedelta.relativedelta(days=1)
    return [m2_eop, m1_eop, m0_eop]


def append_to_excel(write_dict, today_date, wb_obj, start_row=3):
    start_date = get_sop(today_date)
    for sheet_name, df in write_dict.items():
        work_sheet = wb_obj[sheet_name]
        work_sheet['G2'] = start_date
        for df_index, row in df.iterrows():
            row_index = df_index + start_row
            for df_col in range(df.shape[1]):
                column_letter = get_column_letter(df_col + 1)
                cell_index = column_letter + str(row_index)
                work_sheet[cell_index] = row[df_col]


def read_suppliers(filename):
    '''
    Read supplier list from yaml filename
    :return: dictionary of country, suppliers pair
    '''
    with open(filename, 'r') as stream:
        dictionary = yaml.load(stream, Loader=yaml.FullLoader)
    supplier_dict = {}
    for key, value in dictionary.items():
        supplier_dict[str(key)] = value
    return supplier_dict


def get_main_df(country, today_date, data_queried):
    '''
    Get the data for the last five months for specific country between date m2_sop and today_date
    :return: main_df as pandas dataframe
    '''
    country_folder_path = './' + country + '/'
    if not os.path.exists(country_folder_path):
        os.makedirs(country_folder_path)
    input_file_path = "{}input_files".format(country_folder_path)
    if not os.path.exists(input_file_path):
        os.makedirs(input_file_path)
    if not data_queried:
        m2_sop = get_sop(today_date).strftime("%Y-%m-%d")
        today_date = today_date.strftime("%Y-%m-%d")
        query = '''
            SELECT category_cluster, supplier_name, sku_id, cdate, color,
                    cogs_usd, stock_on_hand, inbound_value_usd, acct_payables_usd, 
                    inventory_value_usd, payment_terms, brand, grass_date 
            FROM shopee_bi_sbs_mart
            WHERE
                supplier_name is NOT NULL
            AND
                grass_date >= date('{start_date}')
            AND
                grass_date <= date('{end_date}')
            AND
                purchase_type = 'Outright'
            AND 
                grass_region = '{cntry}'
            '''.format(start_date=m2_sop, end_date=today_date, cntry=country)
        print(query)
        main_df = SPARK.sql(query)
        write_to_csv(main_df, 'main_df', input_file_path + '/main_df.csv')
        main_df = main_df.toPandas()
    else:
        main_df = pd.read_csv(input_file_path + '/main_df.csv',
                              encoding='utf-8-sig', index_col=False)
    return main_df


def write_to_csv(query_df, report_directory_name, output_file_name):
    '''
    This function writes the spark dataframe to a csv file
    :param query_df: spark dataframe
    :type query_df: spark dataframe
    :param output_file_name: Name of Output File
    :type output_file_name: str
    :param report_directory_name: Name of Report Directory
    :type report_directory_name: str
    :return:
    '''
    print("Output File Name: ", output_file_name)
    req_headers = query_df.columns
    username = os.getcwd().split('/')[-1]
    output_filename = 'user/{u}/{d}/tmp.csv'.format(u=username, d=report_directory_name)
    local_filename = output_file_name
    query_df.write.format('csv').mode('overwrite'). \
        options(header='false', escape='"', encoding="UTF-8").save(output_filename)
    subprocess.call('/usr/share/hadoop/bin/hadoop fs -getmerge %s %s' % (
        output_filename, local_filename + "_tmp"), shell=True)
    subprocess.call('/usr/share/hadoop/bin/hadoop fs -rm -r %s' % output_filename, shell=True)
    db_input_file = open(local_filename + "_tmp", 'rb')
    print(db_input_file)
    db_output_file = open(local_filename, 'wb')
    db_output_file.write(','.join(req_headers) + '\n')
    for line in db_input_file:
        db_output_file.write(line)
    db_output_file.close()
    os.remove(local_filename + "_tmp")


def main(country, today_date, main_df, supplier_dict):
    '''
    Pull data; write to the template and save tracking table and
    pivot tables for inventory, COGS, inbounds and COGS
    :param country: country
    :param today_date: last day of the report
    :param main_df: main data frame
    :param supplier_dict: dictionary of country, suppliers list pair
    '''
    half_time = time.time()
    print("--- %s seconds --- generating data" % (half_time - start_time))
    country_folder_path = './' + country + '/'
    main_df['grass_date'] = pd.to_datetime(main_df['grass_date'])
    main_df['cdate'] = pd.to_datetime(main_df['cdate'])
    df_info = main_df[main_df['grass_date'] == today_date]

    # get each supplier's main category by the category cluster of its skus
    # with the highest total inventory value on the latest day
    df_inv_sorting = main_df.sort_values(['grass_date'], ascending=False) \
        .groupby(['supplier_name'])['supplier_name', 'grass_date'] \
        .head(1).dropna(subset=['supplier_name'])
    df_inv_sorting0 = df_inv_sorting.merge(main_df, on=['supplier_name', 'grass_date'])
    df_inv_sorting1 = df_inv_sorting0.groupby(['category_cluster', 'supplier_name'])[
        'inventory_value_usd'].sum().reset_index()
    df_inv_sorting2 = df_inv_sorting1.sort_values(['supplier_name', 'inventory_value_usd'],
                                                  ascending=False).reset_index() \
        .groupby('supplier_name')['supplier_name', 'category_cluster'].head(1).reset_index()
    df_inv_sorting2 = df_inv_sorting2[['category_cluster', 'supplier_name']]

    # Get the total inventory count (stock on hand) of each supplier
    df_inv_count = df_info.groupby('supplier_name')['stock_on_hand'].sum().reset_index()
    df_inv_count.columns = ['supplier_name', 'inv_count']
    df_inv_count = df_inv_sorting2.merge(df_inv_count, on=['supplier_name'], how='left')
    df_inv_sum = df_info.groupby('supplier_name')['inventory_value_usd'].sum().reset_index()

    # get the total number of SKUs each supplier has
    df_sku_count = df_info.groupby('supplier_name')['sku_id'].count().reset_index()
    df_sku_count.columns = ['supplier_name', 'no_skus_WH']
    df_info0 = df_sku_count.merge(df_inv_count, on=['supplier_name'], how='left')

    df_payment = df_info.groupby('supplier_name')['supplier_name', 'payment_terms']. \
        head(1).reset_index(drop=True)
    df_last_month = main_df[(main_df['grass_date'] >= today_date - dt.timedelta(days=29))
                            & (main_df['grass_date'] <= today_date)]
    df_last_month.drop_duplicates(inplace=True)
    brands_df = df_last_month[['supplier_name', 'brand']]
    brands_df = brands_df[brands_df.brand != 'nan']
    number_of_top_brands = 3
    supplier_names = set(brands_df['supplier_name'].to_list())
    brand_count_df = brands_df.groupby(['supplier_name', 'brand']).size()
    # get nlargest per subgroup
    brand_count_top_3_df = brand_count_df.groupby(['supplier_name'], group_keys=False).apply(
        lambda subgroup: subgroup.nlargest(number_of_top_brands))
    # drop count column and reset_index
    brand_count_top_3_df = brand_count_top_3_df.reset_index().drop(columns=0)
    # final brands dataframe in desired output
    top_brands_cols = ['supplier_name', 'brand_1', 'brand_2', 'brand_3']
    top_brands_df = pd.DataFrame(columns=top_brands_cols)
    for supplier in supplier_names:
        slice_df = brand_count_top_3_df[(brand_count_top_3_df.supplier_name == supplier)]. \
            reset_index()
        three_brands = dict()
        for i in range(number_of_top_brands):
            key = 'brand_{}'.format(i + 1)
            try:
                brand = slice_df.iloc[i].brand
                three_brands[key] = brand
            except IndexError:
                three_brands[key] = 'n.a.'
        row_list = [supplier, three_brands['brand_1'], three_brands['brand_2'],
                    three_brands['brand_3']]
        row_to_append = pd.DataFrame([row_list], columns=top_brands_cols)
        top_brands_df = top_brands_df.append(row_to_append)

    top_brands_df = top_brands_df.reset_index(drop=True)
    top_brands_df.fillna('n.a.', inplace=True)
    df_info1 = df_info0.merge(df_inv_sum, on=['supplier_name'], how='left')
    df_info2 = df_info1.merge(top_brands_df, on=['supplier_name'], how='left')
    df_info3 = df_info2.merge(df_payment, on=['supplier_name'], how='left')

    df_repln = main_df[['supplier_name', 'inbound_value_usd', 'cdate', 'grass_date', 'color']]
    df_repln1 = df_repln[['supplier_name']].drop_duplicates()
    df_green_repln = df_repln[['supplier_name']].drop_duplicates()
    eops = get_eops(today_date)
    for i in range(3):
        if eops[i] == calendar.monthrange(eops[i].year, eops[i].month)[1]:
            start_date = eops[i].replace(day=1)
        else:
            start_date = eops[i] - dt.timedelta(days=29)
        df_total_temp = df_repln[(df_repln['grass_date'] >= start_date) & (df_repln['grass_date'] <= eops[i])]
        df_total_temp['is_replenishment'] = np.where(df_total_temp['cdate'] < start_date, 1, 0)
        df_total_temp['inb_repln_usd'] = df_total_temp['inbound_value_usd'] * df_total_temp['is_replenishment']
        df_total_temp['is_green_repln'] = np.where((df_total_temp['color'] == 'Green')
                                                   & (df_total_temp['is_replenishment'] == 1), 1, 0)
        df_total_temp['green_repln_usd'] = df_total_temp['inbound_value_usd'] * df_total_temp['is_green_repln']
        df_repln_mi = df_total_temp.groupby(['supplier_name'])[['inb_repln_usd']].sum()
        df_repln1 = df_repln1.merge(df_repln_mi, on=['supplier_name'], how='left')
        df_green_repln_mi = df_total_temp.groupby(['supplier_name'])[['green_repln_usd']].sum()
        df_green_repln = df_green_repln.merge(df_green_repln_mi, on=['supplier_name'], how='left')

    df_repln1.columns = ['supplier_name', 'inb_repln_m2', 'inb_repln_m1', 'inb_repln_m0_30d']
    df_green_repln.columns = ['supplier_name', 'green_repln_m2', 'green_repln_m1', 'green_repln_m0_30d']

    df_eop_inv = main_df[['supplier_name', 'inventory_value_usd', 'grass_date', 'color']]
    df_eop_inv0 = df_eop_inv[(df_eop_inv['grass_date']).isin(eops)]
    df_eop_inv0['is_black_inv'] = np.where(df_eop_inv0['color'] == 'Black', 1, 0)
    df_eop_inv0['black_inv_usd'] = df_eop_inv0['is_black_inv'] * df_eop_inv0['inventory_value_usd']

    df_black_inv0 = df_eop_inv0.groupby(['supplier_name', 'grass_date']) \
        [['black_inv_usd']].sum() \
        .pivot_table(values='black_inv_usd', index='supplier_name', columns='grass_date') \
        .reset_index()
    if df_black_inv0.shape[1] == 3:
        print('Warning: There is likely to be no data for the date you specified')
    df_black_inv0.columns = ['supplier_name', 'black_inv_m2', 'black_inv_m1', 'black_inv_m0']
    df_eop_inv1 = df_eop_inv0.groupby(['supplier_name', 'grass_date']) \
        [['inventory_value_usd']].sum() \
        .pivot_table(values='inventory_value_usd', index='supplier_name', columns='grass_date') \
        .reset_index()
    df_eop_inv1.columns = ['supplier_name', 'eop_inv_m2', 'eop_inv_m1', 'eop_inv_m0']
    df_eop_inv2 = df_eop_inv1.merge(df_black_inv0, on=['supplier_name'], how='left')

    df_total_sum = main_df.groupby(['supplier_name', 'grass_date']).sum().reset_index()

    df_monthly_inb = df_total_sum[['supplier_name']].drop_duplicates()
    df_monthly_cogs = df_total_sum[['supplier_name']].drop_duplicates()
    df_monthly_inv = df_total_sum[['supplier_name']].drop_duplicates()
    df_monthly_payable = df_total_sum[['supplier_name']].drop_duplicates()
    for i in range(3):
        start_date = eops[i] - dt.timedelta(days=89)
        df_total_temp = df_total_sum[(df_total_sum['grass_date'] >= start_date)
                                     & (df_total_sum['grass_date'] <= eops[i])]
        df_monthly_cogs_mi = df_total_temp.groupby(['supplier_name'])[['cogs_usd']].sum()
        df_monthly_cogs = df_monthly_cogs.merge(df_monthly_cogs_mi, on=['supplier_name'], how='left')
        df_monthly_inv_mi = df_total_temp.groupby(['supplier_name'])[['inventory_value_usd']].mean()
        df_monthly_inv = df_monthly_inv.merge(df_monthly_inv_mi, on=['supplier_name'], how='left')
        df_monthly_payable_mi = df_total_temp.groupby(['supplier_name'])[['acct_payables_usd']].mean()
        df_monthly_payable = df_monthly_payable.merge(df_monthly_payable_mi, on=['supplier_name'], how='left')
        if eops[i] == calendar.monthrange(eops[i].year, eops[i].month)[1]:
            start_date1 = eops[i].replace(day=1)
        else:
            start_date1 = eops[i] - dt.timedelta(days=29)
        df_total_temp1 = df_total_sum[(df_total_sum['grass_date'] >= start_date1)
                                      & (df_total_sum['grass_date'] <= eops[i])]
        df_monthly_inb_mi = df_total_temp1.groupby(['supplier_name'])[['inbound_value_usd']].sum()
        df_monthly_inb = df_monthly_inb.merge(df_monthly_inb_mi, on=['supplier_name'], how='left')

    df_total_sum_m0 = df_total_sum[df_total_sum['grass_date'] >= today_date - dt.timedelta(days=29)]
    df_monthly_inb.columns = ['supplier_name', 'inb_m2', 'inb_m1', 'inb_m0_30d']

    df_monthly_cogs_m0 = df_total_sum_m0.groupby(['supplier_name'])[['cogs_usd']].sum()
    df_monthly_cogs = df_monthly_cogs.merge(df_monthly_cogs_m0, on=['supplier_name'], how='left')
    df_monthly_cogs.columns = ['supplier_name', 'cogs_m2', 'cogs_m1', 'cogs_m0', 'cogs_m0_30d']

    df_monthly_inv_m0 = df_total_sum_m0.groupby(['supplier_name'])[['inventory_value_usd']].mean()
    df_monthly_inv = df_monthly_inv.merge(df_monthly_inv_m0, on=['supplier_name'], how='left')
    df_monthly_inv.columns = ['supplier_name', 'inv_m2', 'inv_m1', 'inv_m0', 'inv_m0_30d']

    df_monthly_payable_m0 = df_total_sum_m0.groupby(['supplier_name'])[['acct_payables_usd']].mean()
    df_monthly_payable = df_monthly_payable.merge(df_monthly_payable_m0, on=['supplier_name'], how='left')
    df_monthly_payable.columns = ['supplier_name', 'payable_m2', 'payable_m1', 'payable_m0', 'payable_m0_30d']

    # pivot tables
    df_total_sum['grass_date'] = df_total_sum['grass_date'].dt.strftime('%Y-%m-%d')
    df_cogs = pd.pivot_table(df_total_sum, values='cogs_usd', index='supplier_name',
                             columns='grass_date').reset_index()
    df_payable = pd.pivot_table(df_total_sum, values='acct_payables_usd',
                                index='supplier_name', columns='grass_date').reset_index()
    df_inv_value = pd.pivot_table(df_total_sum, values='inventory_value_usd',
                                  index='supplier_name', columns='grass_date').reset_index()
    df_inbound = pd.pivot_table(df_total_sum, values='inbound_value_usd',
                                 index='supplier_name', columns='grass_date').reset_index()

    df_info3 = df_info3[['category_cluster', 'supplier_name', 'no_skus_WH',
                         'inv_count', 'inventory_value_usd', 'brand_1',
                         'brand_2', 'brand_3', 'payment_terms']]

    if supplier_dict.get(country) is None:
        df_info4 = df_info3
    else:
        supplier_highlight = supplier_dict.get(country)
        df_info4 = df_info3[df_info3['supplier_name'].isin(supplier_highlight)] \
            .reset_index(drop=True)

    df_info5 = df_info4.merge(df_monthly_inb, on=['supplier_name'], how='left') \
        .merge(df_repln1, on=['supplier_name'], how='left') \
        .merge(df_green_repln, on=['supplier_name'], how='left') \
        .merge(df_eop_inv2, on=['supplier_name'], how='left')

    df_info5['inb_repln_m2_perc'] = df_info5['inb_repln_m2'] / df_info5['inb_m2']
    df_info5['inb_repln_m1_perc'] = df_info5['inb_repln_m1'] / df_info5['inb_m1']
    df_info5['inb_repln_m0_30d_perc'] = df_info5['inb_repln_m0_30d'] / df_info5['inb_m0_30d']
    df_info5['green_repln_m2_perc'] = df_info5['green_repln_m2'] / df_info5['inb_repln_m2']
    df_info5['green_repln_m1_perc'] = df_info5['green_repln_m1'] / df_info5['inb_repln_m1']
    df_info5['green_repln_m0_30d_perc'] = df_info5['green_repln_m0_30d'] / df_info5['inb_repln_m0_30d']
    df_info5['black_inv_m2_perc'] = df_info5['black_inv_m2'] / df_info5['eop_inv_m2']
    df_info5['black_inv_m1_perc'] = df_info5['black_inv_m1'] / df_info5['eop_inv_m1']
    df_info5['black_inv_m0_perc'] = df_info5['black_inv_m0'] / df_info5['eop_inv_m0']
    df_info5.fillna('N/A', inplace=True)
    df_info5 = df_info5[['category_cluster', 'supplier_name', 'no_skus_WH', 'inv_count',
                         'inventory_value_usd', 'brand_1', 'brand_2', 'brand_3', 'payment_terms',
                         'inb_repln_m2_perc', 'inb_repln_m1_perc', 'inb_repln_m0_30d_perc',
                         'green_repln_m2_perc', 'green_repln_m1_perc', 'green_repln_m0_30d_perc',
                         'black_inv_m2', 'black_inv_m1', 'black_inv_m0',
                         'black_inv_m2_perc', 'black_inv_m1_perc', 'black_inv_m0_perc']]

    path = country_folder_path + 'template/wc_template_' + country + '_v3.xlsx'
    wb_obj = openpyxl.load_workbook(path)
    main_ws = wb_obj['Tracking']
    main_ws['B1'] = today_date
    for df_index, row in df_info5.iterrows():
        for col_index in range(9):
            cell_index = get_column_letter(col_index + 1) + str(df_index + 4)
            main_ws[cell_index] = row[col_index]
        for col_index in range(9, 21):
            cell_index = get_column_letter(col_index + 7 + 1) + str(df_index + 4)
            main_ws[cell_index] = row[col_index]

    df_payable1 = df_inv_sorting2.merge(df_monthly_payable, on=['supplier_name'], how='right') \
        .merge(df_payable, on=['supplier_name'], how='inner')
    df_cogs1 = df_inv_sorting2.merge(df_monthly_cogs, on=['supplier_name'], how='right') \
        .merge(df_cogs, on=['supplier_name'], how='inner')
    df_inv_value1 = df_inv_sorting2.merge(df_monthly_inv, on=['supplier_name'], how='right') \
        .merge(df_inv_value, on=['supplier_name'], how='inner')
    df_inbound1 = df_inv_sorting2.merge(df_monthly_inb, on=['supplier_name'], how='right') \
        .merge(df_inbound, on=['supplier_name'], how='inner')

    write_dict = {
        'Daily Accounts Payable': df_payable1,
        'Daily Inventory Value': df_inv_value1,
        'Daily COGS': df_cogs1,
        'Daily Inbounds': df_inbound1
    }

    append_to_excel(write_dict, today_date, wb_obj, start_row=3)
    today = dt.datetime.strftime(today_date, "%Y-%m-%d")
    wb_obj.save(country_folder_path + '/Weekly Working Capital {} - {}.xlsx'.format(country, today))

    df_info3.to_csv(country_folder_path + '{}_tracking_tab.csv'.format(country),
                    encoding="utf-8-sig", index=False)
    df_payable.to_csv(country_folder_path + '{}_payables.csv'.format(country),
                      encoding="utf-8-sig", index=False)
    df_cogs.to_csv(country_folder_path + '{}_cogs.csv'.format(country),
                   encoding="utf-8-sig", index=False)
    df_inv_value.to_csv(country_folder_path + '{}_inventory_value.csv'.format(country),
                        encoding="utf-8-sig", index=False)
    df_inbound.to_csv(country_folder_path + '{}_inbound.csv'.format(country),
                      encoding="utf-8-sig", index=False)
    result_dict = {'tracking': df_info3, 'payable': df_payable, 'cogs': df_cogs,
                   'inventory': df_inv_value, 'inbound': df_inbound, 'tracking_yaml': df_info4}
    final_time = time.time()
    print("--- %s seconds ---processing time" % (final_time - half_time))
    return result_dict


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('-c', '--countries', nargs='+', default=['ID', 'MY', 'TH', 'VN', 'TW', 'PH'])
    parser.add_argument("-d", "--date",
                        default=dt.date.today() - dt.timedelta(days=1),
                        type=lambda d: dt.datetime.strptime(d, '%Y%m%d').date(),
                        help="Date in the format yyyymmdd")
    parser.add_argument('-q', '--queried', help='has data been queried?', default=False)
    args = parser.parse_args()
    supplier_dict0 = read_suppliers('./suppliers.yaml')
    for country0 in args.countries:
        main_df0 = get_main_df(country0, args.date, args.queried)
        main(country0, args.date, main_df0, supplier_dict0)
